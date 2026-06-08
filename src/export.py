"""
Exporter.

Writes output files for each pipeline run into a dedicated folder:

  <output_dir>/<topic-slug>_<date>/
      papers_ranked.xlsx   — rich Excel workbook (two sheets)
      papers_ranked.csv    — plain CSV for scripting / pandas
      paper_summaries.jsonl — one JSON object per summarised paper
      survey_report.md     — Markdown report grouped by topic

The run folder name is derived from the configured topics + today's date so
each run is stored separately and old results are never overwritten.

Excel workbook sheets:
  "Ranked Papers"  — all scored papers, one row each, sortable columns
  "Summaries"      — top-N papers that received LLM summaries, with full
                     findings, scope, taxonomy etc. in readable columns
"""

from __future__ import annotations

import csv
import json
import logging
import re
import shutil
import subprocess
from collections import defaultdict
from datetime import date
from pathlib import Path

from .models import (
    FieldGuide,
    FieldMegaArchitecture,
    JudgeResult,
    LandmarkPaper,
    Paper,
    PaperArchitecture,
    PaperSummary,
    SystemDesign,
    ReadingPath,
    ScoredPaper,
)

logger = logging.getLogger(__name__)

# Columns written to both the CSV and the "Ranked Papers" XLSX sheet
_PAPER_COLUMNS = [
    "rank",
    "title",
    "year",
    "venue",
    "authors",
    "topic_queries",
    "citation_count",
    "influential_citation_count",
    "background_citation_count",
    "influential_ratio",
    "canonical_score",
    "authority_tier",
    "llm_authority_assessment",
    "recommended_action",
    "orientation",
    "quality_score",
    "doi",
    "arxiv_id",
    "url",
    "pdf_url",
    "abstract",
    "sources",
]

# Columns written to the "Summaries" XLSX sheet
_SUMMARY_COLUMNS = [
    "rank",
    "title",
    "year",
    "venue",
    "quality_score",
    "research_scope",
    "core_problem",
    "taxonomy",
    "main_methods",
    "main_findings",
    "limitations",
    "future_directions",
    "datasets_and_benchmarks",
    "url",
    "doi",
]


# ─────────────────────────────────────────────────────────────────────────────
# Public helpers
# ─────────────────────────────────────────────────────────────────────────────

def make_run_dir(topics: list[str], base_dir: Path) -> Path:
    """
    Return (and create) a uniquely-named subdirectory for this run.

    Name format: <topic-slug>[_<topic-slug>...]_YYYY-MM-DD[_N]
    e.g. "ai-computer-science-education_machine-learning_2026-05-20"

    If the directory already exists a counter suffix is appended so old runs
    are never overwritten.
    """
    slugs = [_topic_slug(t) for t in topics[:3]]   # max 3 topics in name
    if len(topics) > 3:
        slugs.append(f"and-{len(topics) - 3}-more")
    date_str = date.today().isoformat()
    stem = "_".join(slugs) + "_" + date_str

    folder = base_dir / stem
    if not folder.exists():
        folder.mkdir(parents=True, exist_ok=True)
        return folder

    # Append incrementing counter to avoid clobbering an earlier run today
    counter = 2
    while True:
        candidate = base_dir / f"{stem}_{counter}"
        if not candidate.exists():
            candidate.mkdir(parents=True, exist_ok=True)
            return candidate
        counter += 1


def _topic_slug(topic: str) -> str:
    """'AI for computer science education' → 'ai-computer-science-education'"""
    _STOPWORDS = {"for", "in", "of", "the", "a", "an", "and", "or", "with", "on"}
    words = re.sub(r"[^\w\s]", "", topic.lower()).split()
    significant = [w for w in words if w not in _STOPWORDS][:4]
    return "-".join(significant) or "topic"


# ─────────────────────────────────────────────────────────────────────────────
# Exporter
# ─────────────────────────────────────────────────────────────────────────────

class Exporter:
    def __init__(self, output_dir: str | Path):
        self._output_dir = Path(output_dir)
        self._output_dir.mkdir(parents=True, exist_ok=True)
        # Set by the caller before exporting: when True, report files are
        # slug-prefixed so multiple topics don't overwrite each other at the root.
        self.multi_topic = False

    def _topic_dir(self, topic: str) -> Path:
        """Return (and create) a per-topic sub-folder inside the run directory."""
        d = self._output_dir / _topic_slug(topic)
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _report_path(self, topic: str, ext: str) -> Path:
        """
        Path for report.md / report.html — kept at the run ROOT, beside
        papers_ranked.xlsx. Prefixed with the topic slug only when the run
        covers several topics (so they don't collide).
        """
        name = f"{_topic_slug(topic)}-report.{ext}" if self.multi_topic else f"report.{ext}"
        return self._output_dir / name

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def export_xlsx(
        self,
        scored_papers: list[ScoredPaper],
        summary_pairs: list[tuple[ScoredPaper, PaperSummary]],
    ) -> Path:
        """
        Write a two-sheet Excel workbook.

        Sheet 1 "Ranked Papers" — every scored paper, all columns.
        Sheet 2 "Summaries"     — only papers that have LLM summaries.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning(
                "openpyxl not installed — skipping XLSX export. "
                "Run: pip install openpyxl"
            )
            return self._output_dir / "papers_ranked.xlsx"

        path = self._output_dir / "papers_ranked.xlsx"
        wb = openpyxl.Workbook()

        # ── Sheet 1: Ranked Papers ──────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Ranked Papers"
        _write_sheet(
            ws1,
            headers=_PAPER_COLUMNS,
            rows=[_to_paper_row(rank, sp) for rank, sp in enumerate(scored_papers, 1)],
            col_widths={
                "rank": 6, "title": 60, "year": 8, "venue": 28,
                "authors": 36, "topic_queries": 28,
                "citation_count": 12, "influential_citation_count": 14,
                "background_citation_count": 14, "influential_ratio": 12,
                "canonical_score": 12, "authority_tier": 16,
                "llm_authority_assessment": 18, "recommended_action": 16,
                "orientation": 16,
                "quality_score": 12, "doi": 22, "arxiv_id": 20,
                "url": 32, "pdf_url": 32, "abstract": 70, "sources": 14,
            },
            url_cols={"url", "pdf_url"},
        )

        # ── Sheet 2: Summaries ─────────────────────────────────────────
        if summary_pairs:
            ws2 = wb.create_sheet("Summaries")
            summary_map = {sp.paper.title: (sp, summary) for sp, summary in summary_pairs}
            summary_rows = []
            rank = 0
            for sp in scored_papers:
                if sp.paper.title in summary_map:
                    rank += 1
                    _, summary = summary_map[sp.paper.title]
                    summary_rows.append(_to_summary_row(rank, sp, summary))
            _write_sheet(
                ws2,
                headers=_SUMMARY_COLUMNS,
                rows=summary_rows,
                col_widths={
                    "rank": 6, "title": 55, "year": 8, "venue": 24,
                    "quality_score": 12, "research_scope": 40,
                    "core_problem": 40, "taxonomy": 40,
                    "main_methods": 40, "main_findings": 60,
                    "limitations": 40, "future_directions": 40,
                    "datasets_and_benchmarks": 36, "url": 32, "doi": 22,
                },
                url_cols={"url"},
            )

        wb.save(path)
        logger.info("XLSX exported: %s (%d papers, %d summaries)",
                    path, len(scored_papers), len(summary_pairs))
        return path

    def export_csv(
        self,
        scored_papers: list[ScoredPaper],
        judge_map: dict[str, JudgeResult] | None = None,
        arch_map: dict[str, PaperArchitecture] | None = None,
    ) -> Path:
        """Write all ranked papers to CSV."""
        path = self._output_dir / "papers_ranked.csv"
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=_PAPER_COLUMNS)
            writer.writeheader()
            for rank, sp in enumerate(scored_papers, start=1):
                jr = (judge_map or {}).get(sp.paper.title)
                arch = (arch_map or {}).get(sp.paper.title)
                writer.writerow(_to_paper_row(rank, sp, jr, arch))
        logger.info("CSV exported: %s (%d papers)", path, len(scored_papers))
        return path

    def export_jsonl(
        self,
        summary_pairs: list[tuple[ScoredPaper, PaperSummary]],
        judge_map: dict[str, JudgeResult] | None = None,
    ) -> Path:
        """Write one JSON object per summarised paper to JSONL."""
        path = self._output_dir / "paper_summaries.jsonl"
        with path.open("w", encoding="utf-8") as f:
            for sp, summary in summary_pairs:
                p = sp.paper
                total = max(p.citation_count, 1)
                jr = (judge_map or {}).get(p.title)
                record = {
                    "title": p.title,
                    "year": p.year,
                    "venue": p.venue,
                    "quality_score": sp.quality_score,
                    "doi": p.doi,
                    "arxiv_id": p.arxiv_id,
                    "url": p.url,
                    "citation_count": p.citation_count,
                    "influential_citation_count": p.influential_citation_count,
                    "background_citation_count": p.background_citation_count,
                    "influential_ratio": round(p.influential_citation_count / total, 4),
                    "canonical_score": p.canonical_score,
                    "authority_tier": p.authority_tier,
                    "topic_queries": p.topic_queries,
                    "sources": p.sources,
                    "judge": jr.to_flat_dict() if jr else None,
                    "summary": summary.to_flat_dict(),
                }
                f.write(json.dumps(record, ensure_ascii=False) + "\n")
        logger.info("JSONL exported: %s (%d papers)", path, len(summary_pairs))
        return path

    def export_architecture_report(
        self,
        topic: str,
        arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
        mega: FieldMegaArchitecture,
        judge_map: dict[str, JudgeResult] | None = None,
        reading_path: "ReadingPath | None" = None,
        system_design: "SystemDesign | None" = None,
        landmarks: "list[LandmarkPaper] | None" = None,
        field_map_style: str = "outline",
    ) -> Path:
        """
        Write the full architecture report for one topic to its sub-folder.

        Part 1 — Field Architecture (mega-arch, Mermaid, gaps)
        Part 2 — Survey Navigator   (orientation map, coverage matrix, reading path)
        Part 3 — System Design      (top-down layered architecture of the field)
        Part 4 — Paper Cards        (one card per paper, with anchor IDs)

        All cross-references inside the file use Markdown anchor links so the
        user can click between sections in any Markdown viewer.
        """
        path = self._report_path(topic, "md")
        md = _build_report_markdown(
            topic, arch_triples, mega, judge_map, reading_path,
            system_design, landmarks, field_map_style,
        )
        path.write_text(md, encoding="utf-8")
        logger.info("Architecture report exported: %s (%d papers)", path, len(arch_triples))
        return path

    def export_html_report(
        self,
        topic: str,
        arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
        mega: FieldMegaArchitecture,
        judge_map: dict[str, JudgeResult] | None = None,
        reading_path: "ReadingPath | None" = None,
        system_design: "SystemDesign | None" = None,
        landmarks: "list[LandmarkPaper] | None" = None,
    ) -> Path:
        """
        Write a self-contained `report.html` — the same report, rendered in the
        browser, where the Field Map can be toggled between an outline and a
        Mermaid diagram with a click. Double-click to open; no server needed.
        """
        # Build the report body with a placeholder where the Field Map goes,
        # so the HTML page can inject the interactive (outline/diagram) tabs.
        body_md = _build_report_markdown(
            topic, arch_triples, mega, judge_map, reading_path,
            system_design, landmarks, field_map_style="__slot__",
        )
        field_map_tree = _field_map_tree_data(mega)
        field_map_mermaid = _field_map_mindmap_mermaid(mega)
        field_tree = _field_tree_html_data(mega)
        problem_tree = _problem_tree_html_data(mega)
        system_design_data = _system_design_html_data(system_design)

        html = _build_html_report(
            topic, body_md, field_map_tree, field_tree, problem_tree,
            system_design_data, field_map_mermaid,
        )
        path = self._report_path(topic, "html")
        path.write_text(html, encoding="utf-8")
        logger.info("HTML report exported: %s", path)
        return path

    def export_system_design_json(
        self,
        topic: str,
        design: "SystemDesign",
    ) -> "Path | None":
        """Write SystemDesign as JSON.  Returns None if synthesis failed."""
        if design.extraction_failed:
            logger.warning(
                "Skipping system-design JSON for '%s' — synthesis failed: %s",
                topic, design.failure_reason,
            )
            return None
        path = self._topic_dir(topic) / "system_design.json"
        path.write_text(
            json.dumps(design.model_dump(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("System design JSON exported: %s (%d layers)", path, len(design.layers))
        return path

    def export_reading_path_json(
        self,
        topic: str,
        reading_path: "ReadingPath",
    ) -> "Path | None":
        """Write ReadingPath as JSON.  Returns None if generation failed."""
        if reading_path.generation_failed:
            logger.warning(
                "Skipping reading path JSON for '%s' — generation failed: %s",
                topic, reading_path.failure_reason,
            )
            return None
        path = self._topic_dir(topic) / "reading_path.json"
        path.write_text(
            json.dumps(reading_path.model_dump(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("Reading path JSON exported: %s (%d steps)", path, len(reading_path.steps))
        return path

    def export_mindmap_html(
        self,
        topic: str,
        mega: FieldMegaArchitecture,
        arch_triples: "list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]] | None" = None,
    ) -> "Path | None":
        """
        Write a self-contained interactive HTML mind map using markmap.

        Opens in any browser — no installation required.  Each node is enriched
        with bold labels, `coverage` badges, italic definitions, technique lists,
        and clickable links to the original papers.  Click any node to expand /
        collapse its children; use the toolbar to reset view or expand all.

        Returns None if synthesis failed.
        """
        if mega.synthesis_failed:
            return None
        path = self._topic_dir(topic) / "mindmap.html"
        markdown = _build_mindmap_markdown(topic, mega, arch_triples or [])
        html = _MINDMAP_HTML_TEMPLATE.format(topic=topic, markdown=markdown)
        path.write_text(html, encoding="utf-8")
        logger.info("Interactive mind map exported: %s", path)
        return path

    def export_mega_architecture_json(
        self,
        topic: str,
        mega: FieldMegaArchitecture,
    ) -> Path | None:
        """Write FieldMegaArchitecture as JSON for programmatic consumption.
        Returns None and skips the file if synthesis failed."""
        if mega.synthesis_failed:
            logger.warning(
                "Skipping JSON export for '%s' — synthesis failed: %s",
                topic, mega.failure_reason,
            )
            return None
        path = self._topic_dir(topic) / "architecture.json"
        path.write_text(
            json.dumps(mega.model_dump(), ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("Architecture JSON exported: %s", path)
        return path

    def export_mega_architecture_mmd(
        self,
        topic: str,
        mega: FieldMegaArchitecture,
    ) -> Path | None:
        """Write the Mermaid diagram as a standalone .mmd file.
        Returns None and skips the file if synthesis failed or diagram is empty."""
        if mega.synthesis_failed or not mega.mermaid_diagram:
            logger.warning(
                "Skipping MMD export for '%s' — synthesis failed or no diagram.",
                topic,
            )
            return None
        topic_dir = self._topic_dir(topic)
        mmd_path = topic_dir / "architecture.mmd"
        mmd_path.write_text(mega.mermaid_diagram, encoding="utf-8")
        logger.info("Mermaid diagram exported: %s", mmd_path)

        png_path = topic_dir / "architecture.png"
        _render_mermaid_png(mmd_path, png_path)
        return mmd_path

    def export_markdown(
        self,
        scored_papers: list[ScoredPaper],
        summary_pairs: list[tuple[ScoredPaper, PaperSummary]],
    ) -> Path:
        """Write a Markdown report grouped by topic."""
        path = self._output_dir / "survey_report.md"

        summary_map: dict[str, PaperSummary] = {
            sp.paper.title: summary for sp, summary in summary_pairs
        }

        by_topic: dict[str, list[ScoredPaper]] = defaultdict(list)
        for sp in scored_papers:
            topic = sp.paper.topic_queries[0] if sp.paper.topic_queries else "Uncategorised"
            by_topic[topic].append(sp)

        lines: list[str] = [
            "# AI Survey Paper Report",
            "",
            f"*Generated by AI Survey Paper Miner — {len(scored_papers)} papers across "
            f"{len(by_topic)} topics*",
            "",
            "---",
            "",
        ]

        for topic, papers in sorted(by_topic.items()):
            lines.append(f"## {topic.title()}")
            lines.append("")
            lines.append(f"*{len(papers)} papers — ranked by quality score*")
            lines.append("")
            for rank, sp in enumerate(papers, start=1):
                lines += _paper_section(rank, sp, summary_map.get(sp.paper.title))

        with path.open("w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        logger.info("Markdown report exported: %s", path)
        return path


# ─────────────────────────────────────────────────────────────────────────────
# XLSX helpers
# ─────────────────────────────────────────────────────────────────────────────

def _write_sheet(ws, headers, rows, col_widths, url_cols=None):
    """Write headers + rows to a worksheet with formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    url_cols = url_cols or set()
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    WRAP = Alignment(wrap_text=True, vertical="top")
    TOP = Alignment(vertical="top")

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = TOP

    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, row_dict in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_dict.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Wrap text for long fields; top-align everything
            if header in {"abstract", "main_findings", "taxonomy",
                          "main_methods", "limitations", "future_directions"}:
                cell.alignment = WRAP
            else:
                cell.alignment = TOP
            # Add hyperlink for URL columns
            if header in url_cols and value and str(value).startswith("http"):
                cell.hyperlink = str(value)
                cell.font = Font(color="0563C1", underline="single")

    # Set column widths
    for col_idx, header in enumerate(headers, 1):
        width = col_widths.get(header, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Taller default row height for wrapped rows
    for row_idx in range(2, len(rows) + 2):
        ws.row_dimensions[row_idx].height = 60


# ─────────────────────────────────────────────────────────────────────────────
# Row builders
# ─────────────────────────────────────────────────────────────────────────────

def _to_paper_row(
    rank: int,
    sp: ScoredPaper,
    jr: "JudgeResult | None" = None,
    arch: "PaperArchitecture | None" = None,
) -> dict:
    p = sp.paper
    total = max(p.citation_count, 1)
    return {
        "rank": rank,
        "title": p.title,
        "year": p.year or "",
        "venue": p.venue or "",
        "authors": "; ".join(p.authors[:5]),
        "topic_queries": "; ".join(p.topic_queries),
        "citation_count": p.citation_count,
        "influential_citation_count": p.influential_citation_count,
        "background_citation_count": p.background_citation_count,
        "influential_ratio": round(p.influential_citation_count / total, 4),
        "canonical_score": round(p.canonical_score, 4),
        "authority_tier": p.authority_tier or "",
        "llm_authority_assessment": jr.authority_assessment if jr and not jr.judge_failed else "",
        "recommended_action": jr.recommended_action if jr and not jr.judge_failed else "",
        "orientation": arch.orientation if arch and not arch.analysis_failed else "",
        "quality_score": round(sp.quality_score, 1),
        "doi": p.doi or "",
        "arxiv_id": p.arxiv_id or "",
        "url": p.url or "",
        "pdf_url": p.pdf_url or "",
        "abstract": (p.abstract or "").replace("\n", " "),
        "sources": "; ".join(p.sources),
    }


def _to_summary_row(rank: int, sp: ScoredPaper, summary: PaperSummary) -> dict:
    p = sp.paper
    return {
        "rank": rank,
        "title": p.title,
        "year": p.year or "",
        "venue": p.venue or "",
        "quality_score": round(sp.quality_score, 1),
        "research_scope": summary.research_scope or "",
        "core_problem": summary.core_problem or "",
        "taxonomy": "\n".join(f"• {t}" for t in summary.taxonomy),
        "main_methods": "\n".join(f"• {m}" for m in summary.main_methods),
        "main_findings": "\n".join(f"• {f}" for f in summary.main_findings),
        "limitations": "\n".join(f"• {l}" for l in summary.limitations),
        "future_directions": "\n".join(f"• {d}" for d in summary.future_directions),
        "datasets_and_benchmarks": "; ".join(summary.datasets_and_benchmarks),
        "url": p.url or "",
        "doi": p.doi or "",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Markdown helpers
# ─────────────────────────────────────────────────────────────────────────────

def _paper_section(rank: int, sp: ScoredPaper, summary: PaperSummary | None) -> list[str]:
    p = sp.paper
    lines: list[str] = []

    lines.append(f"### {rank}. {p.title}")
    lines.append("")
    lines.append(f"- **Year:** {p.year or 'N/A'}")
    lines.append(f"- **Venue:** {p.venue or 'N/A'}")
    lines.append(f"- **Quality score:** {sp.quality_score}")
    lines.append(f"- **Citations:** {p.citation_count}")

    if p.url:
        lines.append(f"- **URL:** {p.url}")
    if p.arxiv_id:
        lines.append(f"- **arXiv:** https://arxiv.org/abs/{p.arxiv_id}")

    if summary and not summary.summarization_failed:
        if summary.research_scope:
            lines.append(f"- **Scope:** {summary.research_scope}")
        if summary.core_problem:
            lines.append(f"- **Core problem:** {summary.core_problem}")
        if summary.taxonomy:
            lines.append(f"- **Taxonomy:** {', '.join(summary.taxonomy[:5])}")
        if summary.main_findings:
            lines.append("- **Key findings:**")
            for finding in summary.main_findings[:3]:
                lines.append(f"  - {finding}")
        if summary.datasets_and_benchmarks:
            lines.append(f"- **Benchmarks:** {', '.join(summary.datasets_and_benchmarks[:5])}")
        all_keywords = []
        for kw_list in summary.keywords.values():
            all_keywords.extend(kw_list[:3])
        if all_keywords:
            lines.append(f"- **Keywords:** {', '.join(all_keywords[:10])}")
    elif summary and summary.summarization_failed:
        lines.append(f"- *Summary unavailable: {summary.failure_reason}*")

    lines.append("")
    return lines


# ─────────────────────────────────────────────────────────────────────────────
# Architecture report helpers  (three-part per-topic Markdown)
# ─────────────────────────────────────────────────────────────────────────────

def _anchor(text: str) -> str:
    """Convert text to a GitHub-Flavoured-Markdown anchor fragment."""
    slug = re.sub(r"[^\w\s-]", "", text.lower())
    slug = re.sub(r"[\s]+", "-", slug).strip("-")
    return slug


def _paper_anchor(sp: ScoredPaper) -> str:
    """Stable anchor id for a paper card (first 6 words of title + year)."""
    words = re.sub(r"[^\w\s]", "", sp.paper.title.lower()).split()[:6]
    slug = "-".join(words)
    year = str(sp.paper.year or "nd")
    return f"{slug}-{year}"


def _build_report_markdown(
    topic: str,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
    mega: FieldMegaArchitecture,
    judge_map: dict[str, JudgeResult] | None,
    reading_path: "ReadingPath | None",
    system_design: "SystemDesign | None",
    landmarks: "list[LandmarkPaper] | None",
    field_map_style: str,
) -> str:
    """Assemble the full report markdown (shared by the .md and .html exports)."""
    show_system_design = bool(system_design and not system_design.extraction_failed)
    html_mode = field_map_style == "__slot__"   # the HTML export uses the slot signal
    lines: list[str] = []
    lines += _render_part1_field_architecture(
        topic, mega, arch_triples,
        has_landmarks=bool(landmarks),
        has_system_design=show_system_design,
        field_map_style=field_map_style,
    )
    lines += _render_part2_survey_navigator(topic, arch_triples, mega, reading_path)
    if landmarks:
        lines += _render_landmark_papers(landmarks)
    if show_system_design:
        lines += _render_part3_system_design(system_design, html_mode=html_mode)
    lines += _render_part4_paper_cards(arch_triples, judge_map, html_mode=html_mode)
    return "\n".join(lines)


# Self-contained HTML report. marked.js renders the Markdown, mermaid.js renders
# the in-body diagrams (paper-card taxonomies, concept map), and the Field Map
# slot gets an interactive collapsible tree built from __FIELD_MAP_TREE__.
# Substituted: __TITLE__/__REPORT_MD__/__FIELD_MAP_TREE__/__FIELD_TREE__/__PROBLEM_TREE__/__SYSTEM_DESIGN__.
_HTML_REPORT_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__ — Survey Report</title>
<style>
  :root{ --fg:#1f2937; --muted:#6b7280; --border:#e5e7eb; --accent:#2563eb; }
  *{box-sizing:border-box}
  body{max-width:1280px;margin:0 auto;padding:2.5rem 2rem 6rem;
       font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
       line-height:1.65;color:var(--fg);}
  h1{font-size:1.9rem} h2{font-size:1.5rem;border-bottom:2px solid var(--border);padding-bottom:.3rem;margin-top:2.2rem}
  h3{font-size:1.2rem;margin-top:1.8rem} h4{font-size:1.05rem}
  a{color:var(--accent);text-decoration:none} a:hover{text-decoration:underline}
  table{border-collapse:collapse;width:100%;margin:.6rem 0;font-size:.92rem}
  th,td{border:1px solid var(--border);padding:6px 10px;text-align:left;vertical-align:top}
  th{background:#f9fafb}
  code{background:#f3f4f6;padding:1px 5px;border-radius:4px;font-size:.88em}
  pre{background:#f6f8fa;padding:.8rem;border-radius:8px;overflow:auto}
  blockquote{border-left:3px solid #d1d5db;margin:.6rem 0;padding:.2rem .9rem;color:var(--muted)}
  hr{border:none;border-top:1px solid var(--border);margin:2rem 0}
  .fm-tabs{display:flex;gap:.5rem;margin:.6rem 0 1rem;flex-wrap:wrap;align-items:center}
  .fm-tabs button{font-size:.9rem;padding:.35rem .9rem;border:1px solid var(--border);
                  border-radius:8px;background:#fff;color:#374151;cursor:pointer}
  .fm-tabs button.active{background:var(--accent);color:#fff;border-color:var(--accent)}
  .fm-vsep{display:inline-block;width:1px;height:1.1rem;background:var(--border);margin:0 .25rem}
  /* Field Map — radial "architecture map" view (pan / zoom) */
  .fm-map{overflow:hidden;border:1px solid var(--border);border-radius:10px;background:#fff;
          height:72vh;position:relative;cursor:grab;user-select:none}
  .fm-map.grabbing{cursor:grabbing}
  .fm-pz{position:absolute;left:0;top:0;transform-origin:0 0;will-change:transform}
  .fm-map .mermaid{background:transparent;text-align:left;overflow:visible}
  .mermaid{background:#fff;text-align:center;overflow:auto}
  /* Field Map — interactive D3 tree of HTML boxes */
  .fmtree-d3{overflow:auto;border:1px solid var(--border);border-radius:10px;
             background:#fff;padding:.5rem;max-height:78vh}
  .fmn{display:inline-block;box-sizing:border-box;max-width:300px;border:1px solid;
       border-radius:8px;padding:6px 11px;line-height:1.34;word-break:break-word;
       white-space:normal;font-family:-apple-system,Segoe UI,Roboto,sans-serif}
  .fmn-d0{background:var(--accent);color:#fff;border-color:#1d4ed8;font-weight:700;font-size:16px}
  .fmn-d1{background:#eff6ff;color:#1e3a8a;border-color:#93c5fd;font-weight:700;font-size:14.5px}
  .fmn-d2{background:#f8fafc;color:#334155;border-color:#e2e8f0;font-size:14px}
  .fmn-info{border-style:solid;border-color:#bfdbfe}
  .fmn-info:hover{background:#eff6ff;border-color:var(--accent)}
  .fmn-i{color:var(--accent);font-size:.85em;opacity:.7}
  .info-backdrop{display:none;position:fixed;inset:0;background:rgba(15,23,42,.45);
    z-index:50;align-items:center;justify-content:center;padding:1rem}
  .info-card{background:#fff;border-radius:12px;max-width:560px;width:100%;
    padding:1.4rem 1.6rem;box-shadow:0 10px 40px rgba(0,0,0,.25);position:relative}
  .info-title{margin:0 1.6rem .6rem 0;font-size:1.05rem;color:#0f172a}
  .info-body{margin:0;line-height:1.55;color:#334155;white-space:pre-wrap}
  .info-x{position:absolute;top:.5rem;right:.7rem;border:none;background:none;
    font-size:1.5rem;line-height:1;color:#94a3b8;cursor:pointer}
  .info-x:hover{color:#0f172a}
  /* System Design — layered architecture */
  .sd-wrap{display:flex;gap:1.1rem;align-items:stretch;flex-wrap:wrap}
  .sd-main{flex:1;min-width:300px;display:flex;flex-direction:column}
  .sd-layer{border:1px solid var(--border);border-radius:10px;padding:.7rem .9rem;background:#fff}
  .sd-layer-h{font-weight:700;font-size:1rem;color:#0f172a}
  .sd-layer-role{color:var(--muted);font-size:.85rem;margin-top:.15rem}
  .sd-comps{display:flex;flex-wrap:wrap;gap:.45rem;margin-top:.55rem}
  .sd-chip{display:inline-block;border:1px solid var(--border);background:#f8fafc;
    border-radius:999px;padding:.28rem .75rem;font-size:.86rem;color:#334155}
  .sd-chip.sd-info{cursor:pointer;border-color:#bfdbfe}
  .sd-chip.sd-info:hover{background:#eff6ff;border-color:var(--accent)}
  .sd-arrow{align-self:center;color:#94a3b8;font-size:1.25rem;line-height:1;margin:.15rem 0}
  .sd-l0{background:#eff6ff;border-color:#bfdbfe}
  .sd-l1{background:#f0f9ff;border-color:#bae6fd}
  .sd-l2{background:#ecfeff;border-color:#a5f3fc}
  .sd-l3{background:#f0fdfa;border-color:#99f6e4}
  .sd-l4{background:#f0fdf4;border-color:#bbf7d0}
  .sd-l5{background:#fefce8;border-color:#fde68a}
  .sd-side{width:230px;flex:none;display:flex;flex-direction:column;gap:.6rem}
  .sd-side-h{font-weight:600;font-size:.78rem;color:var(--muted);text-transform:uppercase;letter-spacing:.05em}
  .sd-cross{background:#faf5ff;border-color:#e9d5ff}
  .sd-flow{margin:1rem 0 0;color:#334155;background:#f8fafc;border:1px solid var(--border);
    border-radius:8px;padding:.6rem .9rem;font-size:.92rem}
  @media(max-width:820px){ .sd-side{width:100%} }
  .fm-hint{align-self:center;color:var(--muted);font-size:.8rem;margin-left:.3rem}
  .ft-view{position:relative;overflow-x:auto}
  .ft-lines{position:absolute;left:0;top:0;pointer-events:none;overflow:visible;z-index:3}
  .ft-line{fill:none;stroke:var(--accent);stroke-width:2;opacity:.85}
  .ft-cols{display:flex;gap:5rem;flex-wrap:nowrap;position:relative;align-items:flex-start}
  .ft-col{flex:1;min-width:240px}
  .ft-h{font-weight:600;font-size:.78rem;color:var(--muted);margin:.1rem 0 .6rem;
        text-transform:uppercase;letter-spacing:.05em}
  .ft-item{position:relative;z-index:1;border:1px solid var(--border);background:#fff;
           padding:.45rem .7rem;border-radius:8px;cursor:pointer;font-size:.9rem;
           margin:.45rem 0;word-break:break-word;line-height:1.35}
  .ft-item:hover{background:#f3f4f6;border-color:#93c5fd}
  .ft-item.active{background:#dbeafe;color:#1e40af;font-weight:600;border-color:var(--accent)}
  .ft-item.dim{opacity:.4}
  .ft-item.hidden{display:none}
  .ft-hint{color:var(--muted);font-size:.83rem;margin:.4rem 0 0}
  .topbar{position:sticky;top:0;background:#ffffffee;backdrop-filter:blur(6px);
          border-bottom:1px solid var(--border);margin:-2.5rem -1.5rem 1.5rem;
          padding:.55rem 1.5rem;font-size:.8rem;color:var(--muted)}
</style>
</head>
<body>
<div id="content">Loading…</div>

<script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/mermaid@10/dist/mermaid.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/d3@7/dist/d3.min.js"></script>
<script>
const REPORT_MD   = __REPORT_MD__;
const FIELD_MAP_TREE = __FIELD_MAP_TREE__;
const FIELD_TREE  = __FIELD_TREE__;
const PROBLEM_TREE = __PROBLEM_TREE__;
const SYSTEM_DESIGN = __SYSTEM_DESIGN__;
const FIELD_MAP_MERMAID = __FIELD_MAP_MERMAID__;

mermaid.initialize({startOnLoad:false, securityLevel:"loose"});

// 1. render the report markdown
document.getElementById("content").innerHTML = marked.parse(REPORT_MD);

// 1b. give every heading a GitHub-style id, so the Table of Contents links and
//     the "Back to Contents" buttons resolve (recent marked.js no longer adds
//     heading ids automatically).
(function(){
  var seen={};
  document.querySelectorAll("#content h1,#content h2,#content h3,#content h4,#content h5,#content h6").forEach(function(h){
    var base=h.textContent.toLowerCase().trim()
      .replace(/[^\\w\\s-]/g,"").replace(/\\s/g,"-").replace(/^-+|-+$/g,"");
    var id=base, i=1;
    while(id && seen[id]){ id=base+"-"+(i++); }
    if(id){ seen[id]=true; h.id=id; }
  });
})();

// 2. convert ```mermaid code blocks into mermaid containers
document.querySelectorAll("code.language-mermaid").forEach(function(c){
  var d=document.createElement("pre"); d.className="mermaid"; d.textContent=c.textContent;
  c.parentElement.replaceWith(d);
});

// Shared popup that explains a taxonomy node ("what this is") on click.
function openInfoPopup(title, body){
  var bk=document.getElementById("info-popup");
  if(!bk){
    bk=document.createElement("div"); bk.id="info-popup"; bk.className="info-backdrop";
    bk.innerHTML='<div class="info-card" role="dialog" aria-modal="true">'+
      '<button class="info-x" aria-label="Close">×</button>'+
      '<h4 class="info-title"></h4><p class="info-body"></p></div>';
    document.body.appendChild(bk);
    var close=function(){ bk.style.display="none"; };
    bk.addEventListener("click",function(e){ if(e.target===bk) close(); });
    bk.querySelector(".info-x").addEventListener("click",close);
    document.addEventListener("keydown",function(e){ if(e.key==="Escape") close(); });
  }
  bk.querySelector(".info-title").textContent=title;
  bk.querySelector(".info-body").textContent=body;
  bk.style.display="flex";
}

// 3. Reusable interactive box-tree renderer (Field Map + per-paper taxonomies).
//    Every node is an HTML box (foreignObject → wrapping div) — full text, no
//    truncation, bold names via data.html. Click a node to expand/collapse;
//    leaves that carry an explanation open an info popup instead.
function renderBoxTree(holder, DATA, opts){
  opts = opts || {};
  var initialDepth = (opts.initialDepth==null ? 1 : opts.initialDepth);
  if(!DATA || !DATA.children || !DATA.children.length){ holder.innerHTML="<p style='color:#6b7280'>—</p>"; return null; }
  if(typeof d3==="undefined"){            // CDN blocked → readable text fallback
    holder.innerHTML="<pre style='white-space:pre-wrap'>"+
      DATA.children.map(function(c){
        return "• "+c.label+"\\n"+(c.children||[]).map(function(i){return "   – "+i.label;}).join("\\n");
      }).join("\\n")+"</pre>";
    return null;
  }
  var MAXW=300, HGAP=48, VGAP=14;
  var root=d3.hierarchy(DATA);
  root.descendants().forEach(function(d,i){ d.id=i; d._children=d.children; if(d.depth>=initialDepth) d.children=null; });
  var svg=d3.create("svg");
  var gLink=svg.append("g").attr("fill","none").attr("stroke","#cbd5e1").attr("stroke-width",1.6);
  var gNode=svg.append("g");
  holder.appendChild(svg.node());

  function escHtml(s){ var e=document.createElement("div"); e.textContent=s; return e.innerHTML; }
  function hasInfo(d){ return !d._children && d.data.desc; }   // leaf with an explanation
  function depthClass(d){ return "fmn fmn-d"+Math.min(d.depth,2)+(hasInfo(d)?" fmn-info":""); }
  function nodeHtml(d){
    var m=(d._children?(d.children?"▾ ":"▸ "):"");
    return escHtml(m)+(d.data.html?d.data.html:escHtml(d.data.label))+(hasInfo(d)?' <span class="fmn-i">ⓘ</span>':"");
  }

  function update(){
    var nodes=root.descendants(), links=root.links();
    var node=gNode.selectAll("g.fmn-g").data(nodes,function(d){return d.id;});
    node.exit().remove();
    var enter=node.enter().append("g").attr("class","fmn-g")
      .on("click",function(e,d){
        if(d._children){ d.children=d.children?null:d._children; update(); }
        else if(d.data.desc){ openInfoPopup(d.data.label, d.data.desc); }
      });
    enter.append("foreignObject").append("xhtml:div");
    var all=enter.merge(node);
    all.style("cursor",function(d){return (d._children||hasInfo(d))?"pointer":"default";});
    all.select("div").attr("class",depthClass).html(nodeHtml);
    all.select("foreignObject").attr("width",MAXW).attr("height",4000);
    all.each(function(d){ var div=this.querySelector("div"); div.style.width=""; d.bw=Math.min(MAXW, Math.ceil(div.getBoundingClientRect().width)+1); });
    all.each(function(d){ var fo=this.querySelector("foreignObject"), div=this.querySelector("div"); div.style.width=d.bw+"px"; d.bh=Math.ceil(div.getBoundingClientRect().height)+1; fo.setAttribute("width",d.bw+2); fo.setAttribute("height",d.bh+2); });
    var maxDepth=0; nodes.forEach(function(d){ if(d.depth>maxDepth)maxDepth=d.depth; });
    var colW=[],k; for(k=0;k<=maxDepth;k++){ var mw=0; nodes.forEach(function(d){ if(d.depth===k&&d.bw>mw)mw=d.bw; }); colW[k]=mw; }
    var colX=[0]; for(k=1;k<=maxDepth;k++){ colX[k]=colX[k-1]+colW[k-1]+HGAP; }
    var cur=0;
    (function lay(d){ d.X=colX[d.depth]; if(d.children&&d.children.length){ d.children.forEach(lay); d.Y=(d.children[0].Y+d.children[d.children.length-1].Y)/2; } else { d.Y=cur+d.bh/2; cur+=d.bh+VGAP; } })(root);
    var t=svg.transition().duration(220);
    all.transition(t).attr("transform",function(d){ return "translate("+d.X+","+(d.Y-d.bh/2)+")"; });
    all.select("foreignObject").attr("x",0).attr("y",0);
    function linkPath(d){ var sx=d.source.X+d.source.bw, sy=d.source.Y, tx=d.target.X, ty=d.target.Y, mx=(sx+tx)/2; return "M"+sx+","+sy+"C"+mx+","+sy+" "+mx+","+ty+" "+tx+","+ty; }
    var link=gLink.selectAll("path").data(links,function(d){return d.target.id;});
    link.exit().remove();
    link.enter().append("path").merge(link).transition(t).attr("d",linkPath);
    var W=0,H=0; nodes.forEach(function(d){ if(d.X+d.bw>W)W=d.X+d.bw; if(d.Y+d.bh/2>H)H=d.Y+d.bh/2; });
    svg.attr("width",W+24).attr("height",H+24).attr("viewBox",[-12,-12,W+24,H+24]);
  }
  update();
  function setAll(expand){ (function walk(d){ var kids=d._children; if(kids){ d.children = expand ? kids : (d.depth>=initialDepth?null:kids); kids.forEach(walk); } })(root); update(); }
  return { expandAll:function(){setAll(true);}, collapseAll:function(){setAll(false);} };
}

// 3a. Field Map — Tree view (default) + Map view (radial mindmap); same data.
(function(){
  var slot=document.getElementById("fieldmap-slot"); if(!slot) return;

  var tabs=document.createElement("div"); tabs.className="fm-tabs";
  tabs.innerHTML='<button id="fmv-tree" class="active">🌳 Tree</button>'+
                 '<button id="fmv-map">🗺 Map</button>'+
                 '<span class="fm-vsep"></span>'+
                 '<button id="fm-expand">⊕ Expand all</button>'+
                 '<button id="fm-collapse">⊖ Collapse all</button>'+
                 '<span class="fm-hint" id="fm-hint">Click a node to expand / collapse · scroll to pan</span>';

  var treeWrap=document.createElement("div");
  var holder=document.createElement("div"); holder.className="fmtree-d3";
  treeWrap.appendChild(holder);

  var mapWrap=document.createElement("div"); mapWrap.className="fm-map"; mapWrap.style.display="none";

  slot.innerHTML=""; slot.appendChild(tabs); slot.appendChild(treeWrap); slot.appendChild(mapWrap);

  var ctrl=renderBoxTree(holder, FIELD_MAP_TREE, {initialDepth:1});
  if(ctrl){ tabs.querySelector("#fm-expand").onclick=ctrl.expandAll; tabs.querySelector("#fm-collapse").onclick=ctrl.collapseAll; }

  var mapDone=false;
  function show(which){
    var isMap=which==="map";
    tabs.querySelector("#fmv-tree").classList.toggle("active", !isMap);
    tabs.querySelector("#fmv-map").classList.toggle("active", isMap);
    tabs.querySelector("#fm-expand").style.display=isMap?"none":"";
    tabs.querySelector("#fm-collapse").style.display=isMap?"none":"";
    tabs.querySelector("#fm-hint").textContent=isMap
      ? "Click a node to expand / collapse & zoom to it · drag to pan · scroll to zoom · double-click to fit"
      : "Click a node to expand / collapse · scroll to pan";
    treeWrap.style.display=isMap?"none":"";
    mapWrap.style.display=isMap?"":"none";
    if(isMap && !mapDone){ mapDone=true; renderFieldMapMap(mapWrap); }   // lazy — needs a visible box
  }
  tabs.querySelector("#fmv-tree").onclick=function(){ show("tree"); };
  tabs.querySelector("#fmv-map").onclick=function(){ show("map"); };
})();

// Render the Field Map as an interactive radial tree (the "architecture map"):
// colour-coded branches, each node expands/collapses AND zooms to itself on
// click; drag to pan, scroll to zoom, double-click background to fit. Same data
// as the Tree view (FIELD_MAP_TREE).
function renderFieldMapMap(mapWrap){
  if(typeof d3==="undefined" || !FIELD_MAP_TREE || !FIELD_MAP_TREE.children){
    mapWrap.innerHTML="<p style='color:#6b7280'>—</p>"; return;
  }
  var W=mapWrap.clientWidth||800, H=mapWrap.clientHeight||520;
  var FILL=["#fef3c7","#dcfce7","#fce7f3","#ede9fe","#fee2e2","#dbeafe","#cffafe","#ffedd5"];
  var STROKE=["#f59e0b","#22c55e","#ec4899","#8b5cf6","#ef4444","#3b82f6","#06b6d4","#f97316"];

  var svg=d3.create("svg").attr("width",W).attr("height",H).style("display","block");
  var g=svg.append("g");
  var gLink=g.append("g").attr("fill","none").attr("stroke","#cbd5e1").attr("stroke-width",1.4).attr("stroke-opacity",.7);
  var gNode=g.append("g");
  mapWrap.innerHTML=""; mapWrap.appendChild(svg.node());

  var root=d3.hierarchy(FIELD_MAP_TREE);
  var ic=0; root.descendants().forEach(function(d){ d.id=++ic; d._children=d.children; if(d.depth>=1) d.children=null; });
  function bidx(d){ var a=d; while(a.depth>1) a=a.parent; return (a.parent&&a.parent.children)? a.parent.children.indexOf(a):0; }
  function trunc(s,n){ s=String(s); return s.length>n? s.slice(0,n-1)+"…": s; }

  var R=Math.max(240, Math.min(W,H)/2-30);
  var tree=d3.tree().size([2*Math.PI,R]).separation(function(a,b){ return (a.parent===b.parent?1:1.7)/Math.max(1,a.depth); });
  var linkGen=d3.linkRadial().angle(function(d){return d.x;}).radius(function(d){return d.y;});

  var zoom=d3.zoom().scaleExtent([0.15,6]).on("zoom",function(e){ g.attr("transform",e.transform); });
  svg.call(zoom).on("dblclick.zoom",null).on("dblclick",function(){ fitAll(); });
  function fitAll(){ svg.transition().duration(400).call(zoom.transform, d3.zoomIdentity.translate(W/2,H/2).scale(0.85)); }
  function zoomTo(d){ var s=1.5, p=d3.pointRadial(d.x,d.y);
    svg.transition().duration(500).call(zoom.transform, d3.zoomIdentity.translate(W/2-p[0]*s,H/2-p[1]*s).scale(s)); }
  svg.call(zoom.transform, d3.zoomIdentity.translate(W/2,H/2).scale(0.85));

  function update(source){
    tree(root);
    var nodes=root.descendants(), links=root.links();

    var link=gLink.selectAll("path").data(links,function(d){return d.target.id;});
    link.exit().remove();
    link.enter().append("path").merge(link).transition().duration(450).attr("d",linkGen);

    var node=gNode.selectAll("g.rm-node").data(nodes,function(d){return d.id;});
    node.exit().remove();
    var enter=node.enter().append("g").attr("class","rm-node").style("cursor","pointer")
      .attr("transform",function(){ var p=d3.pointRadial(source.x0!=null?source.x0:source.x, source.y0!=null?source.y0:source.y); return "translate("+p[0]+","+p[1]+")"; })
      .on("click",function(e,d){ e.stopPropagation(); if(d._children){ d.children=d.children?null:d._children; update(d);} zoomTo(d); });
    enter.append("rect").attr("rx",7).attr("ry",7);
    enter.append("text").attr("dy","0.32em").attr("text-anchor","middle");
    enter.append("title");
    var all=enter.merge(node);
    all.select("text")
      .attr("font-size",function(d){return d.depth===0?14:12;})
      .attr("font-weight",function(d){return d.depth<=1?700:400;})
      .attr("fill",function(d){return d.depth===0?"#fff":"#1f2937";})
      .text(function(d){ return trunc(d.data.label, d.depth===0?34:40) + (d._children?(d.children?"  ▾":"  ▸"):""); });
    all.select("title").text(function(d){return d.data.label;});
    all.each(function(d){
      var bb=this.querySelector("text").getBBox(), r=this.querySelector("rect"), px=10, py=6;
      r.setAttribute("x",bb.x-px); r.setAttribute("y",bb.y-py);
      r.setAttribute("width",bb.width+2*px); r.setAttribute("height",bb.height+2*py);
      r.setAttribute("fill", d.depth===0?"#2563eb": FILL[bidx(d)%FILL.length]);
      r.setAttribute("stroke", d.depth===0?"#1d4ed8": STROKE[bidx(d)%STROKE.length]);
      r.setAttribute("stroke-width",1.2);
    });
    all.transition().duration(450).attr("transform",function(d){ var p=d3.pointRadial(d.x,d.y); return "translate("+p[0]+","+p[1]+")"; });

    root.each(function(d){ d.x0=d.x; d.y0=d.y; });
  }
  update(root);
}

// 3b. Per-paper taxonomy trees — same interactive box tree (data in data-tax)
document.querySelectorAll("div.taxtree").forEach(function(slot){
  var data; try{ data=JSON.parse(slot.getAttribute("data-tax")); }catch(e){ return; }
  slot.removeAttribute("data-tax"); slot.classList.add("fmtree-d3");
  renderBoxTree(slot, data, {initialDepth:1});    // collapsed by default — click a node to expand
});

// 4. render all mermaid diagrams (hidden ones still render; just not shown)
mermaid.run();

// 5. Interactive two-column linked views (Field Tree + Problem Tree)
function renderLinkedTree(slotId, TREE){
  var slot=document.getElementById(slotId);
  if(!slot || !TREE.pairs || !TREE.pairs.length){
    if(slot) slot.innerHTML="<p style='color:#6b7280'>Not enough data to build this tree.</p>";
    return;
  }
  var pairs=TREE.pairs;
  function esc(s){var d=document.createElement("div");d.textContent=s;return d.innerHTML;}
  var tabs='<div class="fm-tabs">';
  pairs.forEach(function(p,i){
    tabs+='<button class="ft-pair'+(i===0?" active":"")+'" data-i="'+i+'">'+esc(p.leftLabel)+" ↔ "+esc(p.rightLabel)+"</button>";
  });
  tabs+='</div><div class="ft-view"></div><p class="ft-hint">Click an item to show only what it connects to, with lines linking them. Click it again to bring everything back.</p>';
  slot.innerHTML=tabs;
  slot.querySelectorAll(".ft-pair").forEach(function(b){
    b.addEventListener("click",function(){
      slot.querySelectorAll(".ft-pair").forEach(function(x){x.classList.remove("active");});
      b.classList.add("active"); drawPair(pairs[+b.dataset.i]);
    });
  });
  drawPair(pairs[0]);

  function drawPair(p){
    var links=p.links, lefts=Object.keys(links), rights=[], rev={};
    lefts.forEach(function(l){(links[l]||[]).forEach(function(r){if(rights.indexOf(r)<0)rights.push(r);});});
    rights.forEach(function(r){rev[r]=[];});
    lefts.forEach(function(l){(links[l]||[]).forEach(function(r){rev[r].push(l);});});
    var html='<svg class="ft-lines"></svg><div class="ft-cols"><div class="ft-col"><div class="ft-h">'+esc(p.leftLabel)+'</div>'+
      lefts.map(function(l,i){return '<div class="ft-item" data-side="L" data-id="'+i+'">'+esc(l)+'</div>';}).join("")+
      '</div><div class="ft-col"><div class="ft-h">'+esc(p.rightLabel)+'</div>'+
      rights.map(function(r,i){return '<div class="ft-item" data-side="R" data-id="'+i+'">'+esc(r)+'</div>';}).join("")+
      '</div></div>';
    var view=slot.querySelector(".ft-view"); view.innerHTML=html;
    var svg=view.querySelector(".ft-lines");
    var L=view.querySelectorAll('.ft-item[data-side=L]');
    var R=view.querySelectorAll('.ft-item[data-side=R]');
    var sel=null;                 // {s:'L'|'R', i:int} — currently focused item

    function clearLines(){ while(svg.firstChild) svg.removeChild(svg.firstChild); }
    function reset(){
      L.forEach(function(e){e.classList.remove("active","dim","hidden");});
      R.forEach(function(e){e.classList.remove("active","dim","hidden");});
      clearLines();
    }
    function line(aEl,bEl){      // a = left box (line from its right edge) → b = right box (left edge)
      var vb=view.getBoundingClientRect(), ar=aEl.getBoundingClientRect(), br=bEl.getBoundingClientRect();
      var x1=ar.right-vb.left+view.scrollLeft, y1=ar.top+ar.height/2-vb.top+view.scrollTop;
      var x2=br.left -vb.left+view.scrollLeft, y2=br.top+br.height/2-vb.top+view.scrollTop;
      var mx=(x1+x2)/2;
      var pth=document.createElementNS("http://www.w3.org/2000/svg","path");
      pth.setAttribute("class","ft-line");
      pth.setAttribute("d","M"+x1+","+y1+"C"+mx+","+y1+" "+mx+","+y2+" "+x2+","+y2);
      svg.appendChild(pth);
    }
    function sizeSvg(){ svg.setAttribute("width",view.scrollWidth); svg.setAttribute("height",view.scrollHeight); }
    // Focusing an item HIDES everything that is not the item or one of its
    // connections, then draws a line to each connected item on the other side.
    // Lines are drawn AFTER hiding so the coordinates match the reflowed layout.
    function focusL(i){
      reset(); L[i].classList.add("active");
      var rel=links[lefts[i]]||[];
      R.forEach(function(re,ri){ if(rel.indexOf(rights[ri])>=0) re.classList.add("active"); else re.classList.add("hidden"); });
      L.forEach(function(le,li){ if(li!==i) le.classList.add("hidden"); });
      sizeSvg();
      R.forEach(function(re,ri){ if(rel.indexOf(rights[ri])>=0) line(L[i],re); });
    }
    function focusR(i){
      reset(); R[i].classList.add("active");
      var rel=rev[rights[i]]||[];
      L.forEach(function(le,li){ if(rel.indexOf(lefts[li])>=0) le.classList.add("active"); else le.classList.add("hidden"); });
      R.forEach(function(re,ri){ if(ri!==i) re.classList.add("hidden"); });
      sizeSvg();
      L.forEach(function(le,li){ if(rel.indexOf(lefts[li])>=0) line(le,R[i]); });
    }
    L.forEach(function(el,i){ el.addEventListener("click",function(){
      if(sel&&sel.s==='L'&&sel.i===i){ reset(); sel=null; }   // re-click → clear
      else { focusL(i); sel={s:'L',i:i}; }
    });});
    R.forEach(function(el,i){ el.addEventListener("click",function(){
      if(sel&&sel.s==='R'&&sel.i===i){ reset(); sel=null; }
      else { focusR(i); sel={s:'R',i:i}; }
    });});
  }
}
renderLinkedTree("fieldtree-slot", FIELD_TREE);
renderLinkedTree("problemtree-slot", PROBLEM_TREE);

// 6. System Design — top-down layered architecture. Layers stack with down
//    arrows; cross-cutting concerns sit on the side; click a component → popup.
(function(){
  var slot=document.getElementById("systemdesign-slot");
  if(!slot) return;
  var SD=SYSTEM_DESIGN;
  if(!SD || !SD.layers || !SD.layers.length){ slot.innerHTML="<p style='color:#6b7280'>—</p>"; return; }
  function comp(c){
    var chip=document.createElement("span"); chip.className="sd-chip"; chip.textContent=c.name;
    if(c.description){
      chip.classList.add("sd-info");
      chip.appendChild(Object.assign(document.createElement("span"),{className:"fmn-i",textContent:" ⓘ"}));
      chip.addEventListener("click",function(){ openInfoPopup(c.name, c.description); });
    }
    return chip;
  }
  function layerEl(L, cls){
    var d=document.createElement("div"); d.className="sd-layer "+(cls||"");
    var h=document.createElement("div"); h.className="sd-layer-h"; h.textContent=L.name; d.appendChild(h);
    if(L.role){ var r=document.createElement("div"); r.className="sd-layer-role"; r.textContent=L.role; d.appendChild(r); }
    if(L.components && L.components.length){
      var cw=document.createElement("div"); cw.className="sd-comps";
      L.components.forEach(function(c){ cw.appendChild(comp(c)); });
      d.appendChild(cw);
    }
    return d;
  }
  var wrap=document.createElement("div"); wrap.className="sd-wrap";
  var main=document.createElement("div"); main.className="sd-main";
  SD.layers.forEach(function(L,i){
    main.appendChild(layerEl(L,"sd-l"+(i%6)));
    if(i<SD.layers.length-1){ var a=document.createElement("div"); a.className="sd-arrow"; a.textContent="↓"; main.appendChild(a); }
  });
  wrap.appendChild(main);
  if(SD.cross_cutting && SD.cross_cutting.length){
    var side=document.createElement("div"); side.className="sd-side";
    var lbl=document.createElement("div"); lbl.className="sd-side-h"; lbl.textContent="Cross-cutting"; side.appendChild(lbl);
    SD.cross_cutting.forEach(function(L){ side.appendChild(layerEl(L,"sd-cross")); });
    wrap.appendChild(side);
  }
  slot.innerHTML=""; slot.appendChild(wrap);
  if(SD.data_flow){ var f=document.createElement("p"); f.className="sd-flow"; f.innerHTML="<strong>Data flow:</strong> "; f.appendChild(document.createTextNode(SD.data_flow)); slot.appendChild(f); }
})();
</script>
</body>
</html>
"""


def _build_html_report(
    topic: str, body_md: str, field_map_tree: dict | None = None,
    field_tree: dict | None = None, problem_tree: dict | None = None,
    system_design: dict | None = None, field_map_mermaid: str = "",
) -> str:
    """Fill the HTML template, JSON-encoding the strings for safe embedding."""
    def js(s: str) -> str:
        # JSON-encode, then neutralise any "</" so it can't close the <script>
        return json.dumps(s).replace("</", "<\\/")

    def jsdata(d: dict | None, empty: dict | None = None) -> str:
        return json.dumps(d or empty or {"pairs": []}).replace("</", "<\\/")

    return (
        _HTML_REPORT_TEMPLATE
        .replace("__TITLE__", topic.title())
        .replace("__REPORT_MD__", js(body_md))
        .replace("__FIELD_MAP_TREE__", jsdata(field_map_tree, {"children": []}))
        .replace("__FIELD_MAP_MERMAID__", js(field_map_mermaid or ""))
        .replace("__FIELD_TREE__", jsdata(field_tree))
        .replace("__PROBLEM_TREE__", jsdata(problem_tree))
        .replace("__SYSTEM_DESIGN__", jsdata(system_design, {"layers": []}))
    )


# ── Coverage-ordered views of the mega fields ───────────────────────────────
# Both the Field Map and the report's Research Landscape iterate these, so the
# two always show the SAME items in the SAME order: most-covered first.

def _cov_key(info: object) -> int:
    """coverage_count for sorting; missing → -1 so it sorts last."""
    if isinstance(info, dict):
        c = info.get("coverage_count")
        if isinstance(c, int):
            return c
    return -1


def _tasks_by_coverage(mega: FieldMegaArchitecture) -> list[tuple[str, dict]]:
    return sorted(
        [(k, v) for k, v in mega.major_tasks.items() if isinstance(v, dict)],
        key=lambda kv: _cov_key(kv[1]), reverse=True,
    )


def _methods_by_coverage(mega: FieldMegaArchitecture) -> list[tuple[str, dict]]:
    return sorted(
        [(k, v) for k, v in mega.method_families.items() if isinstance(v, dict)],
        key=lambda kv: _cov_key(kv[1]), reverse=True,
    )


def _challenges_by_coverage(mega: FieldMegaArchitecture) -> list[tuple[str, dict]]:
    return sorted(
        [(k, v) for k, v in mega.challenges.items() if isinstance(v, dict)],
        key=lambda kv: _cov_key(kv[1]), reverse=True,
    )


def _benchmarks_by_coverage(mega: FieldMegaArchitecture) -> list[dict]:
    return sorted(
        [d for d in mega.datasets_and_benchmarks if isinstance(d, dict) and d.get("name")],
        key=_cov_key, reverse=True,
    )


def _render_field_outline(mega: FieldMegaArchitecture) -> list[str]:
    """
    Field Map as a directory-style nested outline (instead of a Mermaid
    mind-map). Reads the same structured mega-architecture fields, so it stays
    consistent with the report's tables, but renders as plain nested bullets
    that are readable in ANY viewer without diagram support.
    """
    n = len(mega.source_papers)
    low = max(1, round(n * 0.3))
    out: list[str] = []

    def cov(info: object) -> str:
        if isinstance(info, dict) and isinstance(info.get("coverage_count"), int):
            c = info["coverage_count"]
            return f"  ·  {c}/{n} surveys" + (" ⚠️" if c < low else "")
        return ""

    if mega.major_tasks:
        out.append("- **Research Areas**")
        for name, info in _tasks_by_coverage(mega):
            out.append(f"  - {name}{cov(info)}")

    if mega.method_families:
        out.append("- **Methods**")
        for name, info in _methods_by_coverage(mega):
            reps = ""
            rm = [str(x) for x in (info.get("representative_methods") or [])][:4]
            if rm:
                reps = f" — {', '.join(rm)}"
            out.append(f"  - {name}{reps}{cov(info)}")

    if mega.challenges:
        out.append("- **Challenges**")
        for name, info in _challenges_by_coverage(mega):
            sev = ""
            if str(info.get("severity", "")).strip():
                sev = f" `{str(info['severity']).strip()}`"
            out.append(f"  - {name}{sev}{cov(info)}")

    benches = _benchmarks_by_coverage(mega)
    if benches:
        out.append("- **Benchmarks & Datasets**")
        for d in benches:
            out.append(f"  - {d.get('name', '')}{cov(d)}")

    if mega.open_gaps:
        out.append("- **Research Gaps**")
        for g in mega.open_gaps:
            out.append(f"  - {g.gap}")

    if mega.applications:
        out.append("- **Applications**")
        for a in mega.applications:
            out.append(f"  - {a}")

    return out


def _field_map_tree_mermaid(mega: FieldMegaArchitecture) -> str:
    """
    Field Map as a hierarchical TREE diagram (Mermaid flowchart, left→right):

        Topic ─┬─ Research Areas ─── <task> …
               ├─ Methods ─ <family> ── <representative method> …
               ├─ Benchmarks & Datasets ─ <name> …
               ├─ Challenges ─── <challenge> …
               ├─ Research Gaps ─ <gap> …
               └─ Applications ── <app> …

    Same content as `_render_field_outline`, but drawn as one big readable tree.
    Left→right so the many leaf nodes stack vertically instead of overflowing.
    Returns "" when there is no structured data to draw.
    """
    n = len(mega.source_papers)
    low = max(1, round(n * 0.3))
    counter = [0]

    def nid(prefix: str) -> str:
        counter[0] += 1
        return f"{prefix}{counter[0]}"

    def clean(s: object, maxlen: int = 46) -> str:
        # Strip characters that break Mermaid quoted labels; keep ()/ which are safe.
        t = re.sub(r'["#<>\[\]{}|]', " ", str(s).replace("\n", " "))
        t = re.sub(r"\s+", " ", t).strip()
        if len(t) > maxlen:
            t = t[: maxlen - 1].rstrip() + "…"
        return t or "—"

    def cov(info: object) -> str:
        if isinstance(info, dict) and isinstance(info.get("coverage_count"), int):
            c = info["coverage_count"]
            return f" ({c}/{n})" + (" ⚠️" if c < low else "")
        return ""

    lines = ["flowchart LR"]
    root = "ROOT"
    lines.append(f'  {root}(["{clean(mega.topic.title(), 60)}"])')

    def category(title: str, items: list[tuple[str, list[str]]], label_max: int = 46) -> None:
        items = [it for it in items if it[0]]
        if not items:
            return
        cid = nid("CAT")
        lines.append(f'  {cid}["{clean(title, 40)}"]')
        lines.append(f"  {root} --> {cid}")
        for label, children in items:
            iid = nid("N")
            lines.append(f'  {iid}["{clean(label, label_max)}"]')
            lines.append(f"  {cid} --> {iid}")
            for ch in children:
                if not str(ch).strip():
                    continue
                chid = nid("N")
                lines.append(f'  {chid}(["{clean(ch)}"])')
                lines.append(f"  {iid} --> {chid}")

    # Research Areas
    category("Research Areas", [
        (f"{name}{cov(info)}", [])
        for name, info in _tasks_by_coverage(mega)
    ])
    # Methods — representative methods inline (matches the Outline)
    mf_items: list[tuple[str, list[str]]] = []
    for name, info in _methods_by_coverage(mega):
        reps = [str(x) for x in (info.get("representative_methods") or [])][:4]
        tail = f" — {', '.join(reps)}" if reps else ""
        mf_items.append((f"{name}{cov(info)}{tail}", []))
    category("Methods", mf_items, label_max=78)
    # Challenges (before Benchmarks per request)
    ch_items: list[tuple[str, list[str]]] = []
    for name, info in _challenges_by_coverage(mega):
        sev = ""
        if str(info.get("severity", "")).strip():
            sev = f" ({str(info['severity']).strip()})"
        ch_items.append((f"{name}{sev}{cov(info)}", []))
    category("Challenges", ch_items)
    # Benchmarks & Datasets
    category("Benchmarks & Datasets", [
        (f"{d.get('name', '')}{cov(d)}", []) for d in _benchmarks_by_coverage(mega)
    ])
    # Research Gaps
    category("Research Gaps", [(g.gap, []) for g in mega.open_gaps])
    # Applications
    category("Applications", [(a, []) for a in mega.applications])

    if len(lines) <= 2:           # only ROOT, no categories
        return ""
    return "\n".join(lines)


def _field_map_tree_data(mega: FieldMegaArchitecture) -> dict:
    """
    Field Map as nested JSON for the interactive collapsible tree (HTML):
        {label, children:[{label, children:[...]}]}
    Same content as `_render_field_outline` — Topic → category → item — so the
    HTML tree shows the root + categories first and expands items on click.
    """
    n = len(mega.source_papers)
    low = max(1, round(n * 0.3))

    def cov(info: object) -> str:
        if isinstance(info, dict) and isinstance(info.get("coverage_count"), int):
            c = info["coverage_count"]
            return f" ({c}/{n})" + (" ⚠️" if c < low else "")
        return ""

    def esc(s: object) -> str:
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def b(s: object) -> str:                       # bold a named term
        return "<strong>" + esc(s) + "</strong>"

    cats: list[dict] = []
    if mega.major_tasks:
        cats.append({"label": "Research Areas", "children": [
            {"label": f"{name}{cov(info)}", "html": b(name) + esc(cov(info))}
            for name, info in _tasks_by_coverage(mega)
        ]})
    if mega.method_families:
        ch = []
        for name, info in _methods_by_coverage(mega):
            reps = [str(x) for x in (info.get("representative_methods") or [])][:4]
            tail = f" — {', '.join(reps)}" if reps else ""
            html = b(name) + esc(cov(info))
            if reps:
                html += esc(" — ") + ", ".join(b(r) for r in reps)
            ch.append({"label": f"{name}{cov(info)}{tail}", "html": html})
        cats.append({"label": "Methods", "children": ch})
    # Challenges before Benchmarks (per request)
    if mega.challenges:
        ch = []
        for name, info in _challenges_by_coverage(mega):
            sev = ""
            if str(info.get("severity", "")).strip():
                sev = f" ({str(info['severity']).strip()})"
            ch.append({"label": f"{name}{sev}{cov(info)}",
                       "html": b(name) + esc(sev) + esc(cov(info))})
        cats.append({"label": "Challenges", "children": ch})
    benches = _benchmarks_by_coverage(mega)
    if benches:
        cats.append({"label": "Benchmarks & Datasets", "children": [
            {"label": f"{d.get('name', '')}{cov(d)}", "html": b(d.get("name", "")) + esc(cov(d))}
            for d in benches
        ]})
    if mega.open_gaps:                              # gaps are sentences → no bolding
        cats.append({"label": "Research Gaps",
                     "children": [{"label": g.gap} for g in mega.open_gaps]})
    if mega.applications:
        cats.append({"label": "Applications",
                     "children": [{"label": a, "html": b(a)} for a in mega.applications]})

    return {"label": mega.topic.title(), "children": cats}


def _field_map_mindmap_mermaid(mega: FieldMegaArchitecture) -> str:
    """
    The Field Map as a Mermaid `mindmap` (the radial "architecture map" view) —
    built from the SAME nested data as `_field_map_tree_data`, so the HTML
    report's Tree and Map views show identical content and levels (topic →
    category → item). Returns "" when there is no data.
    """
    data = _field_map_tree_data(mega)
    if not data.get("children"):
        return ""

    def clean(s: object, maxlen: int = 70) -> str:
        # Mindmap node text: strip the chars Mermaid treats as node-shape syntax.
        t = re.sub(r'[()\[\]{}"#;]', " ", " ".join(str(s).split()))
        t = re.sub(r"\s+", " ", t).strip()
        return (t[: maxlen - 1] + "…") if len(t) > maxlen else (t or "—")

    lines = ["mindmap", f"  root(({clean(data['label'], 60)}))"]
    for cat in data["children"]:
        lines.append("    " + clean(cat["label"]))
        for item in cat.get("children", []):
            lines.append("      " + clean(item["label"]))
            for sub in item.get("children", []):
                lines.append("        " + clean(sub["label"]))
    return "\n".join(lines)


# ── Field Tree (problem-solving chain: research area → method → technique) ────

def _field_tree_pairs(
    mega: FieldMegaArchitecture,
) -> tuple[dict[str, list[str]], dict[str, list[str]]]:
    """
    Extract the two many-to-many relations behind the Field Tree:
      task_method:   research-area name -> [method-family names]
      method_tech:   method-family name -> [representative technique names]
    Method names from the LLM are matched back to real method_families keys.
    """
    fam_names = list(mega.method_families.keys())
    fam_lower = {f.lower(): f for f in fam_names}

    def resolve_family(m: str) -> str | None:
        ml = m.lower().strip()
        if ml in fam_lower:
            return fam_lower[ml]
        for fl, fn in fam_lower.items():       # fuzzy: substring either way
            if ml and (ml in fl or fl in ml):
                return fn
        return None

    task_method: dict[str, list[str]] = {}
    for task, info in mega.major_tasks.items():
        if not isinstance(info, dict):
            continue
        ms: list[str] = []
        for m in (info.get("methods") or []):
            fam = resolve_family(str(m))
            if fam and fam not in ms:
                ms.append(fam)
        if ms:
            task_method[task] = ms

    method_tech: dict[str, list[str]] = {}
    for fam, info in mega.method_families.items():
        if isinstance(info, dict):
            techs = [str(t).strip() for t in (info.get("representative_methods") or []) if str(t).strip()]
            if techs:
                method_tech[fam] = techs[:6]

    return task_method, method_tech


def _field_tree_background(mega: FieldMegaArchitecture) -> list[str]:
    """Background drivers = the field's core problems (the '-1' tree level)."""
    return [
        str(cp.get("problem", "")).strip()
        for cp in mega.core_problems
        if isinstance(cp, dict) and str(cp.get("problem", "")).strip()
    ][:8]


def _render_field_tree_outline(mega: FieldMegaArchitecture) -> list[str]:
    """
    Static (Markdown) Field Tree — the problem-solving chain (Topic node omitted):
        Core Problems → Research Areas → Methods → Techniques
    """
    task_method, method_tech = _field_tree_pairs(mega)
    bg = _field_tree_background(mega)
    out: list[str] = []

    if bg:
        out.append("- **Core Problems** *(what drives the field)*")
        for b in bg:
            out.append(f"  - {b}")

    if task_method:
        out.append("- **Research Areas** *(how the core problems are studied)*")
        for task, methods in task_method.items():
            out.append(f"  - **{task}**")
            for m in methods:
                techs = method_tech.get(m, [])
                tail = f" — {', '.join(techs[:5])}" if techs else ""
                out.append(f"    - {m}{tail}")
    elif method_tech:  # fallback: no task→method links, just method → techniques
        out.append("- **Methods**")
        for m, techs in method_tech.items():
            out.append(f"  - {m}" + (f" — {', '.join(techs[:5])}" if techs else ""))

    return out


def _field_tree_html_data(mega: FieldMegaArchitecture) -> dict:
    """
    JSON-able data for the interactive (HTML) two-column linked view, covering
    each adjacent layer of the chain (the Topic node is omitted — Core Problems
    connect straight to the Research Areas they motivate):
        Core Problems ↔ Research Areas ↔ Methods ↔ Techniques
    """
    task_method, method_tech = _field_tree_pairs(mega)
    bg = _field_tree_background(mega)        # full text — the linked view wraps boxes

    pairs: list[dict] = []
    if bg and task_method:
        areas = list(task_method.keys())
        pairs.append({
            "key": "problem-area", "leftLabel": "Core Problems", "rightLabel": "Research Areas",
            "links": {b: areas for b in bg},   # the core problems collectively motivate the areas
        })
    if task_method:
        pairs.append({
            "key": "area-method", "leftLabel": "Research Areas", "rightLabel": "Methods",
            "links": task_method,
        })
    if method_tech:
        pairs.append({
            "key": "method-technique", "leftLabel": "Methods", "rightLabel": "Techniques",
            "links": method_tech,
        })
    return {"pairs": pairs}


# ── Problem Tree (problem-exposing chain: research area → challenge → gap) ────

_BLUESKY_LABEL = "✦ Blue-sky ideas (no current challenge)"


def _problem_tree_relations(
    mega: FieldMegaArchitecture,
) -> tuple[dict[str, list[str]], dict[str, list[str]], list[str]]:
    """
    Extract the relations behind the problem-exposing tree:
      area_chal:  research-area name -> [challenge names]   (many-to-many)
      chal_gap:   challenge name     -> [gap sentences]      (many-to-many)
      free_gaps:  [gap sentences]    gaps tied to NO challenge (blue-sky ideas)
    Challenge names from the LLM are matched back to real `challenges` keys.
    """
    chal_lower = {c.lower(): c for c in mega.challenges.keys()}

    def resolve_chal(c: str) -> str | None:
        cl = c.lower().strip()
        if cl in chal_lower:
            return chal_lower[cl]
        for clk, cn in chal_lower.items():        # fuzzy: substring either way
            if cl and (cl in clk or clk in cl):
                return cn
        return None

    area_chal: dict[str, list[str]] = {}
    for area, info in mega.major_tasks.items():
        if not isinstance(info, dict):
            continue
        cs: list[str] = []
        for c in (info.get("challenges") or []):
            ch = resolve_chal(str(c))
            if ch and ch not in cs:
                cs.append(ch)
        if cs:
            area_chal[area] = cs

    chal_gap: dict[str, list[str]] = {}
    free_gaps: list[str] = []
    for gap in mega.open_gaps:
        g = gap.gap.strip()
        if not g:
            continue
        linked: list[str] = []
        for c in gap.related_challenges:
            ch = resolve_chal(str(c))
            if ch and ch not in linked:
                linked.append(ch)
        if linked:
            for ch in linked:
                chal_gap.setdefault(ch, [])
                if g not in chal_gap[ch]:
                    chal_gap[ch].append(g)
        else:
            free_gaps.append(g)

    return area_chal, chal_gap, free_gaps


def _render_problem_tree_outline(mega: FieldMegaArchitecture) -> list[str]:
    """
    Static (Markdown) Problem Tree — the problem-exposing chain:
        Research Area → Challenge → Research Gap
    "Blue-sky" gaps motivated by no current challenge are listed separately.
    """
    area_chal, chal_gap, free_gaps = _problem_tree_relations(mega)
    out: list[str] = []

    if area_chal:
        for area, chals in area_chal.items():
            out.append(f"- **{area}** *(research area)*")
            for ch in chals:
                out.append(f"  - **{ch}** *(challenge)*")
                for g in chal_gap.get(ch, []):
                    out.append(f"    - {g}")
    elif chal_gap:  # fallback: no area→challenge links, just challenge → gaps
        for ch, gaps in chal_gap.items():
            out.append(f"- **{ch}** *(challenge)*")
            for g in gaps:
                out.append(f"  - {g}")

    if free_gaps:
        out.append("- **Blue-sky ideas** *(open gaps not tied to a current challenge)*")
        for g in free_gaps:
            out.append(f"  - {g}")

    return out


def _problem_tree_html_data(mega: FieldMegaArchitecture) -> dict:
    """
    JSON-able data for the interactive (HTML) two-column linked view of the
    problem-exposing tree, covering each adjacent layer:
        Research Areas ↔ Challenges ↔ Research Gaps
    Blue-sky gaps appear under a special left node in the Challenges ↔ Gaps view.
    """
    area_chal, chal_gap, free_gaps = _problem_tree_relations(mega)

    pairs: list[dict] = []
    if area_chal:
        pairs.append({
            "key": "area-challenge", "leftLabel": "Research Areas",
            "rightLabel": "Challenges", "links": area_chal,
        })
    if chal_gap or free_gaps:                # full gap text — boxes wrap
        links = {ch: list(gaps) for ch, gaps in chal_gap.items()}
        if free_gaps:
            links[_BLUESKY_LABEL] = list(free_gaps)
        pairs.append({
            "key": "challenge-gap", "leftLabel": "Challenges",
            "rightLabel": "Research Gaps", "links": links,
        })
    return {"pairs": pairs}


def _render_part1_field_architecture(
    topic: str,
    mega: FieldMegaArchitecture,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
    has_landmarks: bool = False,
    has_system_design: bool = False,
    field_map_style: str = "outline",
) -> list[str]:
    n = len(mega.source_papers)
    n_gaps = len(mega.open_gaps)
    today = date.today().isoformat()

    lines: list[str] = [
        f"# {topic.title()} — Survey Report",
        f"*Generated {today} · {n} surveys analysed · {n_gaps} research gaps identified*",
        "",
        "---",
        "",
        "## Contents",
        "",
        "- [Part 1 — Field Architecture](#part-1--field-architecture)",
        "  - [Field at a Glance](#field-at-a-glance)",
        "  - [Field Map](#field-map)",
        "  - [Field Tree](#field-tree)",
        "  - [Core Problems](#core-problems)",
        "  - [Research Landscape](#research-landscape)",
        "  - [Research Gaps](#research-gaps)",
        "  - [Problem Tree](#problem-tree)",
        "- [Part 2 — Survey Navigator](#part-2--survey-navigator)",
        "  - [Reading Guide](#reading-guide-where-to-start)",
        *(["- [Landmark Papers](#landmark-papers)"] if has_landmarks else []),
        *([
            "- [Part 3 — System Design](#part-3--system-design)",
        ] if has_system_design else []),
        "- [Part 4 — Paper Cards](#part-4--paper-cards)",
        "",
        "---",
        "",
        "## Part 1 — Field Architecture",
        "",
        "### Field at a Glance",
        "",
    ]

    if mega.field_summary:
        lines.append(mega.field_summary)
        lines.append("")

    # Deepest / weakest coverage bullets derived from coverage_count
    all_methods = list(mega.method_families.items())
    if all_methods:
        best = sorted(all_methods, key=lambda kv: kv[1].get("coverage_count", 0), reverse=True)
        worst = [kv for kv in best if kv[1].get("coverage_count", 0) < max(1, round(n * 0.3))]
        if best:
            lines.append(f"**Deepest coverage:** {' · '.join(k for k, _ in best[:3])}")
        if worst:
            lines.append(f"**Weakest coverage:** {' · '.join(k for k, _ in worst[:3])}")

    cmp = mega.cross_survey_comparison
    if cmp and not cmp.comparison_failed and cmp.conflicting_classifications:
        conflict = cmp.conflicting_classifications[0]
        lines.append(
            f"**Most contested concept:** {conflict.get('dimension', '')} "
            f"({conflict.get('paper_a', '')} vs {conflict.get('paper_b', '')})"
        )
    lines.append("")

    # Field Map — diagram (Mermaid), outline (directory-style list), both, or
    # "__slot__" (HTML export injects interactive outline/diagram tabs here).
    lines += ["---", "", "### Field Map", ""]
    style = (field_map_style or "outline").lower()
    mermaid_block = [
        "```mermaid",
        (_field_map_tree_mermaid(mega) or mega.mermaid_diagram
         or ("mindmap\n  root((" + topic.title() + "))")),
        "```",
        "",
    ]
    if style == "__slot__":
        lines += ['<div id="fieldmap-slot"></div>', ""]
    elif style == "diagram":
        lines += mermaid_block
    elif style == "both":
        lines += mermaid_block
        lines += _render_field_outline(mega)
    else:  # "outline" (default) or any unrecognised value
        lines += _render_field_outline(mega)
    lines += [
        "",
        "> ⚠️ marks items covered by fewer than 30% of analysed surveys — likely research gaps.",
        "",
    ]

    # Field Tree — the problem-solving chain: research area → method → technique
    lines += [
        "---",
        "",
        "### Field Tree",
        "",
        "> The problem-solving chain: **Core Problems → Research Areas → Methods → "
        "Techniques**. (The core problems motivate the research areas, which use "
        "methods, which are realised by techniques.)",
        "",
    ]
    if style == "__slot__":
        lines += ['<div id="fieldtree-slot"></div>', ""]
    else:
        tree = _render_field_tree_outline(mega)
        lines += tree if tree else ["*Not enough data to build the field tree.*"]
        lines.append("")

    lines += [
        "---",
        "",
        "### Core Problems",
        "",
    ]

    if mega.core_problems:
        lines.append("| Problem | Surveys covering it | Best paper |")
        lines.append("|---|---|---|")
        for cp in mega.core_problems:
            prob = cp.get("problem", "")
            cnt = cp.get("coverage_count", "—")
            best_p = cp.get("best_paper", "—")
            best_link = _paper_ref(best_p, arch_triples)
            coverage = f"{cnt} / {n}" if isinstance(cnt, int) else str(cnt)
            lines.append(f"| {prob} | {coverage} | {best_link} |")
    else:
        lines.append("*No core problems extracted.*")
    lines.append("")

    # Merged Research Landscape (mainstream areas + methods + datasets + challenges)
    lines += _render_research_landscape(mega, arch_triples)

    # Research gaps
    lines += ["---", "", "### Research Gaps", ""]
    if mega.open_gaps:
        for i, gap in enumerate(mega.open_gaps, 1):
            score_str = f"{gap.opportunity_score:.2f}" if gap.opportunity_score else "—"
            lines.append(
                f"**Gap {i} — {gap.gap}** *(opportunity score: {score_str})*"
            )
            lines.append("")
            if gap.gap_type:
                lines.append(f"**Type:** {gap.gap_type}")
            if gap.evidence:
                evidence_links = [_paper_ref(title, arch_triples) for title in gap.evidence]
                lines.append(f"**Evidence:** {' · '.join(evidence_links)}")
            lines.append("")
            lines.append("---")
            lines.append("")
    else:
        lines.append("*No research gaps identified.*")
        lines.append("")

    # Problem Tree — the problem-exposing chain: research area → challenge → gap
    lines += [
        "---",
        "",
        "### Problem Tree",
        "",
        "> The problem-exposing chain: **Research Area → Challenge → Research Gap**. "
        "Each area still faces several challenges (many-to-many), and each challenge "
        "opens onto research gaps. *Blue-sky* gaps that no current challenge motivates "
        "are listed on their own.",
        "",
    ]
    if style == "__slot__":
        lines += ['<div id="problemtree-slot"></div>', ""]
    else:
        ptree = _render_problem_tree_outline(mega)
        lines += ptree if ptree else [
            "*Not enough data to build the problem tree "
            "(re-run the pipeline to populate area→challenge and gap→challenge links).*"
        ]
        lines.append("")

    return lines


# Generic words ignored when matching a research-area name to papers.
_AREA_STOPWORDS = {
    "and", "or", "the", "a", "an", "of", "for", "in", "on", "to", "with",
    "based", "using", "systems", "system", "general", "other", "via",
    "approaches", "approach", "methods", "method",
}


def _area_tokens(name: str) -> set[str]:
    return {
        w for w in re.findall(r"[a-z0-9]+", name.lower())
        if w not in _AREA_STOPWORDS and len(w) > 2
    }


def _papers_for_area(
    area_name: str,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
    max_papers: int = 6,
) -> list[str]:
    """
    Return Markdown links to the papers most relevant to a research area.

    Robustly matches by word overlap between the area name and the UNION of
    each paper's covered tasks/applications/challenges, its taxonomy, and its
    summary taxonomy — so an area gets its key papers even when the exact
    phrasing differs (the old strict substring match left many areas empty).
    Ranked by overlap strength, then citation count.
    """
    tokens = _area_tokens(area_name)
    if not tokens:
        return []

    scored: list[tuple[int, int, ScoredPaper]] = []
    for sp, summary, arch in arch_triples:
        if arch.analysis_failed:
            continue
        haystack_parts = (
            list(arch.covered_tasks)
            + list(arch.covered_applications)
            + list(arch.covered_challenges)
            + list(arch.top_level_taxonomy)
        )
        if summary and not summary.summarization_failed:
            haystack_parts += list(summary.taxonomy)
        hay_tokens = set(re.findall(r"[a-z0-9]+", " ".join(haystack_parts).lower()))
        overlap = len(tokens & hay_tokens)
        if overlap >= 1:
            scored.append((overlap, sp.paper.citation_count, sp))

    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    links: list[str] = []
    for _, _, sp in scored[:max_papers]:
        a = _paper_anchor(sp)
        cit = f" ({sp.paper.citation_count:,}✱)" if sp.paper.citation_count else ""
        links.append(f"[{sp.paper.title}](#{a}){cit}")
    return links


def _render_research_landscape(
    mega: FieldMegaArchitecture,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
) -> list[str]:
    """
    Merged section: Research Areas + Methods + Benchmarks & Datasets + Challenges
    + Applications — the same buckets as the Field Map, in the same order.

    For each method family, links to the actual papers that use it (with
    citation counts and paper-card anchors) so readers can navigate directly
    to the relevant literature.
    """
    n = len(mega.source_papers)
    low_threshold = max(1, round(n * 0.3))

    lines: list[str] = [
        "---",
        "",
        "### Research Landscape",
        "",
        "*Research areas, methods in use, relevant benchmarks, and challenges — "
        "all cross-referenced to the surveyed papers below.*",
        "",
    ]

    # ── 1. Research areas ───────────────────────────────────────────────
    if mega.major_tasks:
        lines += ["#### Research Areas", ""]
        lines.append("| Research Area | What it studies | Surveys | Key Papers |")
        lines.append("|---|---|---|---|")
        for task_name, info in _tasks_by_coverage(mega):
            desc = str(info.get("description", "—"))
            cnt = info.get("coverage_count", "—")
            coverage = f"{cnt} / {n}" if isinstance(cnt, int) else str(cnt)
            warn = " ⚠️" if isinstance(cnt, int) and cnt < low_threshold else ""

            # Papers covering this area — robust word-overlap match across all
            # of each paper's covered fields + taxonomy (not just covered_tasks),
            # so areas no longer come up empty just because of phrasing.
            paper_links = _papers_for_area(task_name, arch_triples, max_papers=6)
            # One paper per line inside the cell (<br> renders as a line break)
            papers_str = "<br>".join(paper_links) if paper_links else "—"
            safe_desc = desc.replace("|", "\\|")
            lines.append(f"| **{task_name}{warn}** | {safe_desc} | {coverage} | {papers_str} |")
        lines.append("")

    # ── 2. Methods ──────────────────────────────────────────────────────
    if mega.method_families:
        lines += ["#### Methods", ""]
        lines.append(
            "*For each method: what it is, which benchmarks evaluate it, "
            "and which papers use it (citation count in parentheses).*"
        )
        lines.append("")

        for fam_name, info in _methods_by_coverage(mega):
            desc = str(info.get("description", ""))
            rep_methods: list[str] = [str(x) for x in (info.get("representative_methods") or [])]
            cnt = info.get("coverage_count", "—")
            coverage = f"{cnt} / {n}" if isinstance(cnt, int) else str(cnt)
            warn = " ⚠️" if isinstance(cnt, int) and cnt < low_threshold else ""

            fam_lower = fam_name.lower()
            rep_lower = {r.lower() for r in rep_methods[:5]}

            # Papers that use this method family (match against arch.covered_methods)
            papers_using: list[str] = []
            for sp, _, arch in arch_triples:
                if arch.analysis_failed:
                    continue
                covered_str = " ".join(arch.covered_methods).lower()
                if fam_lower in covered_str or any(r in covered_str for r in rep_lower):
                    a = _paper_anchor(sp)
                    p = sp.paper
                    year_str = str(p.year) if p.year else "n.d."
                    cit_str = f"{p.citation_count:,} citations" if p.citation_count else "no citation data"
                    label = f"{p.title} ({year_str}, {cit_str})"
                    papers_using.append(f"[{label}](#{a})")

            # Render as a compact definition-style block
            lines.append(f"**{fam_name}{warn}** · *{coverage} surveys*  ")
            if desc:
                lines.append(f"> {desc}  ")
            if rep_methods:
                lines.append(f"Techniques: {', '.join(rep_methods[:6])}  ")
            if papers_using:
                lines.append("Papers using this approach:")
                for ref in papers_using[:6]:
                    lines.append(f"- {ref}")
            lines.append("")

    # ── 3. Open challenges (before benchmarks, matching the Field Map) ────
    if mega.challenges:
        lines += ["#### Challenges", ""]
        lines.append("| Challenge | Severity | Surveys | Description |")
        lines.append("|---|---|---|---|")
        for name, info in _challenges_by_coverage(mega):
            sev = str(info.get("severity", "—"))
            cnt = info.get("coverage_count", "—")
            desc = str(info.get("description", "—"))
            coverage = f"{cnt} / {n}" if isinstance(cnt, int) else str(cnt)
            safe_desc = desc.replace("|", "\\|")
            lines.append(f"| **{name}** | {sev} | {coverage} | {safe_desc} |")
        lines.append("")

    # ── 4. Key benchmarks & datasets ────────────────────────────────────
    if mega.datasets_and_benchmarks:
        lines += ["#### Benchmarks & Datasets", ""]
        lines.append("| Benchmark / Dataset | Research area | Surveys citing it |")
        lines.append("|---|---|---|")
        for ds in _benchmarks_by_coverage(mega):
            name = ds.get("name", "")
            task = ds.get("task", "—")
            cnt = ds.get("coverage_count", "—")
            coverage = f"{cnt} / {n}" if isinstance(cnt, int) else str(cnt)
            warn = " ⚠️" if isinstance(cnt, int) and cnt < low_threshold else ""
            lines.append(f"| **{name}**{warn} | {task} | {coverage} |")
        lines.append("")

    # ── 5. Applications ──────────────────────────────────────────────────
    if mega.applications:
        lines += ["#### Applications", ""]
        for a in mega.applications:
            lines.append(f"- {a}")
        lines.append("")

    return lines


def _render_part2_survey_navigator(
    topic: str,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
    mega: FieldMegaArchitecture,
    reading_path: "ReadingPath | None" = None,
) -> list[str]:
    cmp = mega.cross_survey_comparison
    lines: list[str] = [
        "---",
        "",
        "## Part 2 — Survey Navigator",
        "",
    ]

    # Complementary coverage from comparison
    if cmp and not cmp.comparison_failed and cmp.complementary_coverage:
        lines += ["---", "", "### Reading Guide: Where to Start", ""]
        lines.append("| Goal | Best paper |")
        lines.append("|---|---|")
        for item in cmp.complementary_coverage[:8]:
            aspect = item.get("aspect", "")
            best = item.get("best_covered_by", "")
            lines.append(f"| {aspect} | {_paper_ref(best, arch_triples)} |")
        if cmp.best_overall_structure:
            lines.append(
                f"| Best overall structure | "
                f"{_paper_ref(cmp.best_overall_structure, arch_triples)} |"
            )
        lines.append("")

    # Reading path (LLM-generated sequenced reading plan)
    if reading_path and not reading_path.generation_failed and reading_path.steps:
        lines += ["---", "", "### Sequenced Reading Path", ""]
        if reading_path.target_audience:
            lines.append(f"*For: {reading_path.target_audience}*")
            lines.append("")
        lines.append("| Step | Paper | Why | Focus | Est. time |")
        lines.append("|---|---|---|---|---|")
        for step in reading_path.steps:
            # Full real title + link to the paper card (falls back to the
            # step's own title text, un-truncated, if no confident match)
            title_link = _paper_ref(step.paper_title, arch_triples)
            focus = ", ".join(step.focus_sections[:3]) if step.focus_sections else "—"
            time_str = step.estimated_reading_time or "—"
            rationale = step.rationale.replace("|", "\\|")
            lines.append(f"| {step.step} | {title_link} | {rationale} | {focus} | {time_str} |")
        lines.append("")

    return lines


def _render_part4_paper_cards(
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
    judge_map: "dict[str, JudgeResult] | None" = None,
    html_mode: bool = False,
) -> list[str]:
    lines: list[str] = [
        "---",
        "",
        "## Part 4 — Paper Cards",
        "",
    ]

    for sp, summary, arch in arch_triples:
        p = sp.paper
        anchor = _paper_anchor(sp)
        jr = (judge_map or {}).get(p.title)

        lines += ["---", ""]
        lines.append(f'<a id="{anchor}"></a>')
        lines.append("")

        # Title line with metadata badges
        lines.append(f"### {p.title}")
        meta_parts = []
        if p.year:
            meta_parts.append(str(p.year))
        if p.venue:
            meta_parts.append(p.venue)
        meta_parts.append(f"Quality {sp.quality_score:.0f} / 100")
        if p.authority_tier:
            tier_badge = {"foundational": "🏛", "current_standard": "📌", "emerging": "🌱"}.get(
                p.authority_tier, ""
            )
            meta_parts.append(f"{tier_badge} {p.authority_tier}")
        if arch.orientation and not arch.analysis_failed:
            meta_parts.append(f"*{arch.orientation}-oriented*")
        lines.append(f"**{'  ·  '.join(meta_parts)}**")
        lines.append("")

        # URL and back-link
        url = p.url or (f"https://arxiv.org/abs/{p.arxiv_id}" if p.arxiv_id else "")
        if url:
            lines.append(f"[Paper URL]({url})  ·  [Back to Contents](#contents)")
        else:
            lines.append("[Back to Contents](#contents)")
        lines.append("")

        # One-line takeaway from summary
        if summary and not summary.summarization_failed and summary.core_problem:
            lines.append(f"> **One-line takeaway:** {summary.core_problem}")
            lines.append("")

        # Stats table
        lines.append("| Field | Value |")
        lines.append("|---|---|")
        # OpenAlex doesn't expose an "influential citation" count, so it is
        # always 0 — only show it when a source actually provides one.
        if p.influential_citation_count:
            lines.append(
                f"| Citations | {p.citation_count} "
                f"(influential: {p.influential_citation_count}) |"
            )
        else:
            lines.append(f"| Citations | {p.citation_count} |")
        if jr and not jr.judge_failed:
            if jr.recommended_action:
                lines.append(f"| Recommended action | **{jr.recommended_action}** |")
            if jr.scope_clarity:
                lines.append(f"| Scope | {jr.scope_clarity} |")
            if jr.coverage_depth:
                lines.append(f"| Coverage depth | {jr.coverage_depth} |")
        if summary and not summary.summarization_failed and summary.taxonomy:
            lines.append(f"| Covers | {' · '.join(summary.taxonomy[:5])} |")
        lines.append("")

        # Taxonomy tree — interactive box tree in HTML, Mermaid figure in Markdown
        if arch.top_level_taxonomy:
            lines.append("**How this survey organises the field:**")
            lines.append("")
            if html_mode:
                tree = _taxonomy_tree_data(
                    _short_title(p.title),
                    arch.top_level_taxonomy,
                    arch.second_level_taxonomy,
                    glossary=arch.taxonomy_glossary,
                )
                payload = _html_attr(json.dumps(tree, ensure_ascii=False))
                lines.append(f'<div class="taxtree" data-tax="{payload}"></div>')
            else:
                lines += _taxonomy_mermaid(
                    root_label=_short_title(p.title),
                    top_level=arch.top_level_taxonomy,
                    second_level=arch.second_level_taxonomy,
                    anchor=anchor,
                )
            lines.append("")

        # Organizational logic — formatted as scannable bullets, not a wall of text
        if arch.organizational_logic:
            lines.append("**How it organises the field:**")
            lines.append("")
            lines += _format_prose_as_bullets(arch.organizational_logic)
            lines.append("")

        # Structural strengths/weaknesses — bullet lists so every point sits on
        # its own line and renders identically in Markdown and HTML (no run-on).
        if arch.structural_strengths and not arch.analysis_failed:
            lines.append(f"**Read this if:** {arch.structural_strengths[0]}")
            lines.append("")
        if arch.notable_omissions and not arch.analysis_failed:
            lines.append("**Notable omissions:**")
            lines.append("")
            for om in arch.notable_omissions[:3]:
                lines.append(f"- {om}")
            lines.append("")

        # LLM-Judge strengths and weaknesses
        if jr and not jr.judge_failed:
            if jr.strengths:
                lines.append("**Strengths:**")
                lines.append("")
                for s in jr.strengths[:3]:
                    lines.append(f"- {s}")
                lines.append("")
            if jr.weaknesses:
                lines.append("**Weaknesses:**")
                lines.append("")
                for w in jr.weaknesses[:2]:
                    lines.append(f"- {w}")
                lines.append("")
        lines.append("")

    return lines


def _render_part0_field_guide(topic: str, guide: "FieldGuide") -> list[str]:
    """Render Part 0 — Beginner Field Guide."""
    from .field_guide import render_field_guide_markdown
    lines: list[str] = [
        f"# {topic.title()} — Survey Report",
        "",
        "---",
        "",
        "## Part 0 — Field Guide *(Start Here)*",
        "",
        "> This section is a plain-English introduction for newcomers to the field.",
        "",
        render_field_guide_markdown(guide),
        "---",
        "",
    ]
    return lines


# Plain-English metadata for each relationship type.
# (display order, readable verb for the diagram, sentence template, what it means)
_EDGE_META: dict[str, dict] = {
    "is_subfield_of": {
        "label": "is a subfield of",
        "heading": "Hierarchy — subfields",
        "explains": "The first concept is a more specialised area within the second.",
    },
    "part_of": {
        "label": "is part of",
        "heading": "Composition — components",
        "explains": "The first concept is a component or building block of the second.",
    },
    "uses": {
        "label": "uses",
        "heading": "Dependencies — what builds on what",
        "explains": "The first concept relies on or is built using the second.",
    },
    "applied_to": {
        "label": "is applied to",
        "heading": "Applications — method → task",
        "explains": "The first concept (a method) is used to tackle the second (a task or domain).",
    },
    "evaluated_by": {
        "label": "is evaluated by",
        "heading": "Evaluation — benchmarks & metrics",
        "explains": "The first concept is measured using the second (a benchmark or metric).",
    },
    "contrasts_with": {
        "label": "contrasts with",
        "heading": "Contrasts — competing approaches",
        "explains": "The two concepts are alternatives or stand in tension with each other.",
    },
    "emerged_after": {
        "label": "emerged after",
        "heading": "Timeline — what came later",
        "explains": "The first concept developed later than the second.",
    },
}
# Order relationship groups are presented in.
_EDGE_ORDER = [
    "is_subfield_of", "part_of", "uses", "applied_to",
    "evaluated_by", "contrasts_with", "emerged_after",
]


def _render_landmark_papers(landmarks: list["LandmarkPaper"]) -> list[str]:
    """
    Render the Landmark Papers section: the seminal *primary* works (not
    surveys) that the analysed surveys repeatedly build on.
    """
    lines: list[str] = [
        "---",
        "",
        "## Landmark Papers",
        "",
        "*Seminal primary papers (not surveys) that the analysed surveys most often "
        "build upon — read these for the original techniques.*",
        "",
        "| Paper | Year | Citations | Referenced by | Why it matters |",
        "|---|---|---:|---:|---|",
    ]
    for lm in landmarks:
        name = (lm.name or "").strip()
        short = _short_title(lm.title, 9) if lm.title else name
        # Avoid "Self-RAG — Self-RAG: ..." redundancy when the name is in the title
        if name and lm.title and name.lower() not in lm.title.lower():
            title_disp = f"{name} — {short}"
        else:
            title_disp = short or name
        link = f"[{title_disp}]({lm.url})" if lm.url else title_disp
        year = lm.year or "—"
        cites = f"{lm.citation_count:,}" if lm.citation_count else "—"
        ref = f"{lm.mentioned_by} surveys" if lm.mentioned_by else "—"
        why = (lm.why_seminal or "").replace("|", "\\|")
        lines.append(f"| {link} | {year} | {cites} | {ref} | {why} |")
    lines.append("")
    return lines


def _render_part3_system_design(design: "SystemDesign", html_mode: bool = False) -> list[str]:
    """
    Render Part 3 — System Design: a top-down layered architecture of the field.
    In HTML the layers are drawn as an interactive diagram (click a component for
    its explanation); in Markdown they render as ordered layer sections.
    """
    lines: list[str] = [
        "---",
        "",
        "## Part 3 — System Design",
        "",
        "*A top-down view of the field as one system — read it top to bottom to "
        "understand how the pieces fit together.*",
        "",
    ]
    if design.overview:
        lines += [design.overview, ""]

    if html_mode:
        # The interactive layered diagram is injected here from SYSTEM_DESIGN.
        lines += ['<div id="systemdesign-slot"></div>', ""]
        return lines

    # ── Markdown: layers as ordered sections, with a downward arrow between ──
    for i, layer in enumerate(design.layers):
        lines.append(f"### {i + 1}. {layer.name}")
        if layer.role:
            lines += [f"*{layer.role}*", ""]
        if layer.components:
            lines.append("| Component | What it is |")
            lines.append("|---|---|")
            for c in layer.components:
                safe = (c.description or "").replace("|", "\\|")
                lines.append(f"| **{c.name}** | {safe} |")
        lines.append("")
        if i < len(design.layers) - 1:
            lines += ["↓", ""]

    if design.cross_cutting:
        lines += ["### Cross-cutting concerns",
                  "*Span every layer above.*", ""]
        for layer in design.cross_cutting:
            comps = ", ".join(f"**{c.name}**" for c in layer.components) or ""
            role = f" — {layer.role}" if layer.role else ""
            lines.append(f"- **{layer.name}**{role}" + (f": {comps}" if comps else ""))
        lines.append("")

    if design.data_flow:
        lines += [f"**Data flow:** {design.data_flow}", ""]

    return lines


def _system_design_html_data(design: "SystemDesign | None") -> dict:
    """JSON-able SystemDesign for the interactive (HTML) layered diagram."""
    if not design or design.extraction_failed:
        return {"layers": []}

    def layer(L) -> dict:
        return {
            "name": L.name, "role": L.role,
            "components": [{"name": c.name, "description": c.description} for c in L.components],
        }

    return {
        "overview": design.overview,
        "layers": [layer(L) for L in design.layers],
        "cross_cutting": [layer(L) for L in design.cross_cutting],
        "data_flow": design.data_flow,
    }


def _concept_graph_mermaid(
    graph: "ConceptGraph",
    node_name_map: dict[str, str],
    max_edges: int = 24,
) -> list[str]:
    """
    Render the concept graph as a Mermaid directed diagram with labelled edges.

    Only nodes that participate in a shown edge are drawn, so isolated
    concepts don't clutter the figure.  Capped at `max_edges` edges to keep
    the diagram legible; a note is added if edges were omitted.
    """
    edges = graph.edges[:max_edges]
    omitted = len(graph.edges) - len(edges)

    def nid(node_id: str) -> str:
        return "cg" + re.sub(r"[^a-zA-Z0-9]", "_", node_id)[:40]

    out: list[str] = ["```mermaid", "graph LR"]

    # Declare only the nodes that appear in a shown edge
    used_ids: list[str] = []
    seen: set[str] = set()
    for e in edges:
        for node_id in (e.source_id, e.target_id):
            if node_id not in seen:
                seen.add(node_id)
                used_ids.append(node_id)
    for node_id in used_ids:
        name = _mermaid_label(node_name_map.get(node_id, node_id))
        out.append(f'    {nid(node_id)}["{name}"]')

    # Edges with relationship labels
    for e in edges:
        meta = _EDGE_META.get(e.edge_type, {"label": e.edge_type.replace("_", " ")})
        label = _mermaid_label(meta["label"], max_len=22).replace('"', "")
        out.append(f"    {nid(e.source_id)} -->|{label}| {nid(e.target_id)}")

    out.append("```")
    if omitted > 0:
        out.append("")
        out.append(f"> Showing {len(edges)} of {len(graph.edges)} relationships "
                   f"({omitted} omitted for legibility — see the full list below).")
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Interactive HTML mind map (markmap)
# ─────────────────────────────────────────────────────────────────────────────

# Self-contained HTML template.  markmap-autoloader from CDN does all the
# rendering — no build step, no local install.
# Only {topic} and {markdown} are format slots; all other braces are doubled.
_MINDMAP_HTML_TEMPLATE = """\
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{topic} — Mind Map</title>
  <style>
    html, body {{
      margin: 0; padding: 0;
      width: 100%; height: 100%;
      overflow: hidden;
      background: #f0f4f8;
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
    }}
    .mm-header {{
      position: fixed; top: 14px; left: 50%;
      transform: translateX(-50%);
      background: white;
      border-radius: 24px;
      padding: 7px 20px;
      box-shadow: 0 2px 14px rgba(0,0,0,0.12);
      display: flex; align-items: center; gap: 10px;
      z-index: 100;
    }}
    .mm-header h1 {{
      font-size: 0.88rem; font-weight: 700;
      color: #1a202c; white-space: nowrap; margin: 0;
    }}
    .mm-header .sep {{ color: #cbd5e0; }}
    .mm-header button {{
      font-size: 0.72rem; padding: 3px 11px;
      border: 1px solid #e2e8f0; border-radius: 10px;
      background: white; color: #4a5568; cursor: pointer;
      transition: all 0.15s;
    }}
    .mm-header button:hover {{ background: #ebf4ff; border-color: #90cdf4; color: #2b6cb0; }}
    .markmap {{ width: 100vw; height: 100vh; }}
    svg.markmap {{ width: 100%; height: 100%; }}
    /* Make paper-link nodes visually distinct */
    .markmap-node text a {{ text-decoration: underline; }}
  </style>
  <script src="https://cdn.jsdelivr.net/npm/markmap-autoloader@0.16"></script>
</head>
<body>
  <div class="mm-header">
    <h1>{topic}</h1>
    <span class="sep">|</span>
    <button id="btn-fit">⤢ Fit</button>
    <button id="btn-expand">＋ Expand all</button>
    <button id="btn-collapse">－ Collapse all</button>
  </div>

  <div class="markmap">
    <script type="text/template">
---
markmap:
  colorFreezeLevel: 2
  initialExpandLevel: 2
  maxWidth: 360
  zoom: true
  pan: true
---
{markdown}
    </script>
  </div>

  <script>
    // Poll for the markmap SVG instance (autoloader is async)
    let mm = null;
    const poll = setInterval(() => {{
      const svg = document.querySelector("svg.markmap");
      if (svg && svg.__markmap) {{ mm = svg.__markmap; clearInterval(poll); }}
    }}, 80);

    document.getElementById("btn-fit").onclick = () => mm && mm.fit();

    function walkNodes(node, fn) {{
      fn(node);
      (node.children || []).forEach(c => walkNodes(c, fn));
    }}
    document.getElementById("btn-expand").onclick = () => {{
      if (!mm) return;
      walkNodes(mm.state.data, n => {{ n.payload = n.payload || {{}}; n.payload.fold = 0; }});
      mm.renderData(mm.state.data);
      mm.fit();
    }};
    document.getElementById("btn-collapse").onclick = () => {{
      if (!mm) return;
      // Collapse all but the root's direct children
      (mm.state.data.children || []).forEach(child => {{
        walkNodes(child, (n, depth) => {{
          if (n !== child) {{ n.payload = n.payload || {{}}; n.payload.fold = 1; }}
        }});
      }});
      mm.renderData(mm.state.data);
      mm.fit();
    }};

    // Open all links in a new tab
    document.addEventListener("click", e => {{
      const a = e.target.closest("a[href]");
      if (a && !a.href.startsWith("#")) {{
        e.preventDefault();
        window.open(a.href, "_blank", "noopener");
      }}
    }});
  </script>
</body>
</html>
"""


def _build_mindmap_markdown(
    topic: str,
    mega: FieldMegaArchitecture,
    arch_triples: "list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]]",
) -> str:
    """
    Build rich Markdown for markmap from the structured FieldMegaArchitecture.

    Node anatomy:
      - **Concept name** `N/M surveys`          ← bold label + coverage badge
        - *One-sentence definition*             ← italic description child
        - Techniques: `GPT-4` `BERT` `LLaMA`   ← code-span technique badges
        - [Paper title (year)](url) `120✱`      ← clickable citation children

    Nodes with coverage below 30% get a ⚠️ suffix.
    Paper links open in a new tab (handled by the HTML script).
    """
    n = len(mega.source_papers)
    low = max(1, round(n * 0.3))

    def cov(cnt) -> str:
        return f"`{cnt}/{n}`" if isinstance(cnt, int) else f"`{cnt}`"

    def warn(cnt) -> str:
        return " ⚠️" if isinstance(cnt, int) and cnt < low else ""

    def paper_link(sp: ScoredPaper) -> str:
        """One-line citation: linked title + year + citation badge."""
        p = sp.paper
        url = p.url or (f"https://arxiv.org/abs/{p.arxiv_id}" if p.arxiv_id else "")
        year = str(p.year) if p.year else "n.d."
        cit = f" `{p.citation_count:,}✱`" if p.citation_count else ""
        label = f"**{_short_title(p.title, 7)}** ({year})"
        return f"[{label}]({url}){cit}" if url else f"{label}{cit}"

    lines: list[str] = [f"# {topic}", ""]

    # ── Core Problems ────────────────────────────────────────────────────
    if mega.core_problems:
        lines += ["## ❓ Core Problems", ""]
        for cp in mega.core_problems:
            prob = str(cp.get("problem", "")).strip()
            cnt  = cp.get("coverage_count", "—")
            best = str(cp.get("best_paper", "")).strip()
            if not prob:
                continue
            lines.append(f"- **{_truncate_str(prob, 55)}** {cov(cnt)}")
            if best:
                a = _find_paper_anchor(best, arch_triples)
                lines.append(f"  - Best covered by: [{best}](#{a})" if a else f"  - Best covered by: {best}")
        lines.append("")

    # ── Research Areas (major tasks) ─────────────────────────────────────
    if mega.major_tasks:
        lines += ["## 📋 Research Areas", ""]
        for task, info in mega.major_tasks.items():
            if not isinstance(info, dict):
                continue
            cnt  = info.get("coverage_count", "—")
            desc = str(info.get("description", "")).strip()
            lines.append(f"- **{task}**{warn(cnt)} {cov(cnt)}")
            if desc:
                lines.append(f"  - *{desc}*")
            # Papers covering this task
            t_lower = task.lower()
            for sp, _, arch in arch_triples:
                if arch.analysis_failed:
                    continue
                if any(t_lower in t.lower() or t.lower() in t_lower
                       for t in arch.covered_tasks):
                    lines.append(f"  - {paper_link(sp)}")
        lines.append("")

    # ── Methods ──────────────────────────────────────────────────────────
    if mega.method_families:
        lines += ["## 🔧 Methods", ""]
        for fam, info in mega.method_families.items():
            if not isinstance(info, dict):
                continue
            cnt  = info.get("coverage_count", "—")
            desc = str(info.get("description", "")).strip()
            reps: list[str] = [str(x) for x in (info.get("representative_methods") or [])]
            lines.append(f"- **{fam}**{warn(cnt)} {cov(cnt)}")
            if desc:
                lines.append(f"  - *{desc}*")
            if reps:
                badges = " ".join(f"`{r}`" for r in reps[:6])
                lines.append(f"  - Techniques: {badges}")
            # Papers using this method
            f_lower = fam.lower()
            rep_lower = {r.lower() for r in reps[:5]}
            for sp, _, arch in arch_triples:
                if arch.analysis_failed:
                    continue
                covered = " ".join(arch.covered_methods).lower()
                if f_lower in covered or any(r in covered for r in rep_lower):
                    lines.append(f"  - {paper_link(sp)}")
        lines.append("")

    # ── Benchmarks & Datasets ────────────────────────────────────────────
    if mega.datasets_and_benchmarks:
        lines += ["## 📊 Benchmarks", ""]
        for ds in mega.datasets_and_benchmarks:
            name = str(ds.get("name", "")).strip()
            task = str(ds.get("task", "")).strip()
            cnt  = ds.get("coverage_count", "—")
            if not name:
                continue
            lines.append(f"- **{name}**{warn(cnt)} {cov(cnt)}")
            if task:
                lines.append(f"  - *{task}*")
        lines.append("")

    # ── Challenges ────────────────────────────────────────────────────────
    if mega.challenges:
        lines += ["## ⚡ Challenges", ""]
        for name, info in mega.challenges.items():
            if not isinstance(info, dict):
                continue
            cnt  = info.get("coverage_count", "—")
            sev  = str(info.get("severity", "")).strip()
            desc = str(info.get("description", "")).strip()
            sev_badge = f" `{sev}`" if sev else ""
            lines.append(f"- **{name}**{sev_badge} {cov(cnt)}")
            if desc:
                lines.append(f"  - *{desc}*")
        lines.append("")

    # ── Research Gaps ─────────────────────────────────────────────────────
    if mega.open_gaps:
        lines += ["## 🔭 Research Gaps", ""]
        for gap in mega.open_gaps:
            score = f" `score {gap.opportunity_score:.2f}`" if gap.opportunity_score else ""
            gtype = f" `{gap.gap_type}`" if gap.gap_type else ""
            lines.append(f"- **{_truncate_str(gap.gap, 60)}**{gtype}{score}")
            for ev_title in gap.evidence[:2]:
                a = _find_paper_anchor(ev_title, arch_triples)
                ev_link = f"[{_short_title(ev_title)}](#{a})" if a else _short_title(ev_title)
                lines.append(f"  - Evidence: {ev_link}")
        lines.append("")

    # ── Applications ─────────────────────────────────────────────────────
    if mega.applications:
        lines += ["## 🌐 Applications", ""]
        for app in mega.applications:
            lines.append(f"- {app}")
        lines.append("")

    return "\n".join(lines)


def _truncate_str(text: str, max_len: int) -> str:
    return text if len(text) <= max_len else text[: max_len - 1] + "…"


def _render_mermaid_png(mmd_path: Path, png_path: Path) -> None:
    """
    Render a .mmd file to PNG using the Mermaid CLI.

    Tries `mmdc` first (global install), then `npx @mermaid-js/mermaid-cli`
    (no global install required).  Logs a one-line hint and returns silently
    if neither is available — the pipeline is never blocked by this.
    """
    # Candidate mmdc locations: PATH first, then common nvm/homebrew locations
    _MMDC_CANDIDATES = [
        "mmdc",
        "/Users/guanzhongpan/.nvm/versions/node/v20.18.0/bin/mmdc",
        "/opt/homebrew/bin/mmdc",
    ]
    cmd: list[str] | None = None
    for candidate in _MMDC_CANDIDATES:
        if shutil.which(candidate) or (
            candidate != "mmdc" and __import__("os").path.exists(candidate)
        ):
            cmd = [candidate, "-i", str(mmd_path), "-o", str(png_path), "-b", "white"]
            break
    if cmd is None and shutil.which("npx"):
        cmd = [
            "npx", "--yes", "@mermaid-js/mermaid-cli",
            "-i", str(mmd_path), "-o", str(png_path), "-b", "white",
        ]

    if cmd is None:
        logger.info(
            "Mermaid CLI not found — skipping PNG render for %s. "
            "Install with: npm install -g @mermaid-js/mermaid-cli",
            mmd_path.name,
        )
        return

    # Tell Puppeteer to use system Chrome if present (avoids download errors)
    env = {**__import__("os").environ}
    system_chrome = "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"
    if __import__("os").path.exists(system_chrome):
        env["PUPPETEER_EXECUTABLE_PATH"] = system_chrome
        env["PUPPETEER_SKIP_DOWNLOAD"] = "true"

    try:
        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60,
            env=env,
        )
        if result.returncode == 0:
            logger.info("Mermaid PNG rendered: %s", png_path.name)
        else:
            logger.warning(
                "Mermaid render failed for %s (exit %d): %s",
                mmd_path.name, result.returncode,
                (result.stderr or result.stdout).strip()[:200],
            )
    except subprocess.TimeoutExpired:
        logger.warning("Mermaid render timed out for %s", mmd_path.name)
    except Exception as exc:
        logger.warning("Mermaid render error for %s: %s", mmd_path.name, exc)


# Generic words that appear in almost every survey title and therefore carry
# no disambiguating signal.  They are stripped before fuzzy title matching so a
# fragment like "RAG for AIGC: A Survey" doesn't match "RAG for LLMs: A Survey"
# purely on the shared words "for", "a", "survey".
_TITLE_STOPWORDS = {
    "a", "an", "the", "for", "of", "on", "in", "and", "or", "to", "with", "from",
    "survey", "review", "systematic", "literature", "comprehensive", "overview",
    "study", "towards", "via", "using", "based", "approach", "approaches",
    "generation", "retrieval", "augmented", "retrievalaugmented",
}


# Domain acronyms expanded before matching, so a fragment that uses the acronym
# ("RAG for AIGC") matches a title that spells it out ("…for AI-Generated Content").
# RAG/LLM expansions land on stopwords and drop out, leaving the distinguishing
# words (e.g. "ai generated content" vs "large language models").
_TITLE_ACRONYMS = {
    "aigc": "ai generated content",
    "llms": "large language models",
    "llm": "large language model",
    "mllms": "multimodal large language models",
    "mllm": "multimodal large language model",
    "cais": "compound ai systems",
    "rag": "retrieval augmented generation",
    "qa": "question answering",
    "kg": "knowledge graph",
    "nlp": "natural language processing",
    "mcp": "model context protocol",
}


def _title_signature(text: str) -> str:
    """
    Lowercase significant words of a title for fuzzy matching:
    expand domain acronyms, then drop generic stopwords.
    """
    expanded = " ".join(
        _TITLE_ACRONYMS.get(w, w) for w in re.findall(r"[a-z0-9]+", text.lower())
    )
    words = re.findall(r"[a-z0-9]+", expanded)
    return " ".join(w for w in words if w not in _TITLE_STOPWORDS and len(w) > 1)


def _find_paper(
    title_fragment: str,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
) -> "ScoredPaper | None":
    """
    Return the ScoredPaper whose title best matches `title_fragment`, or None.

    Matching is done on the *significant* words of the title (generic survey
    words removed) using rapidfuzz, with a disambiguation margin so an ambiguous
    fragment matches NOTHING rather than a plausible-but-wrong paper.
    """
    if not title_fragment:
        return None

    frag_lower = title_fragment.lower().strip()

    # 1. Exact / containment match on the full title (strongest signal)
    for sp, _, _ in arch_triples:
        t = sp.paper.title.lower()
        if frag_lower == t or (len(frag_lower) > 15 and frag_lower in t) \
           or (len(t) > 15 and t in frag_lower):
            return sp

    # 2. Fuzzy match on significant words, with a disambiguation margin
    from rapidfuzz import fuzz
    frag_sig = _title_signature(title_fragment)
    if not frag_sig:
        return None

    scored: list[tuple[float, ScoredPaper]] = []
    for sp, _, _ in arch_triples:
        score = fuzz.token_set_ratio(frag_sig, _title_signature(sp.paper.title))
        scored.append((score, sp))

    scored.sort(key=lambda x: x[0], reverse=True)
    best_score, best_sp = scored[0]
    second_score = scored[1][0] if len(scored) > 1 else 0.0

    if best_score >= 70 and (best_score - second_score) >= 15:
        return best_sp
    return None


def _find_paper_anchor(
    title_fragment: str,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
) -> str:
    """Anchor of the best-matching paper card, or "" if no confident match."""
    sp = _find_paper(title_fragment, arch_triples)
    return _paper_anchor(sp) if sp else ""


def _paper_ref(
    title_query: str,
    arch_triples: list[tuple[ScoredPaper, PaperSummary, PaperArchitecture]],
) -> str:
    """
    A Markdown link to the matching paper card, using the paper's FULL real
    title (so titles are never confusingly truncated or paraphrased). Falls
    back to the given text (no link) when there is no confident match.
    """
    sp = _find_paper(title_query, arch_triples)
    if sp:
        return f"[{sp.paper.title}](#{_paper_anchor(sp)})"
    return title_query or ""


def _short_title(title: str, max_words: int = 5) -> str:
    words = title.split()
    if len(words) <= max_words:
        return title
    return " ".join(words[:max_words]) + "…"


# Abbreviations whose trailing dot must NOT be treated as a sentence boundary.
_SENTENCE_ABBREV = (
    "e.g.", "i.e.", "etc.", "vs.", "cf.", "et al.", "al.", "Fig.", "Eq.",
    "Sec.", "approx.", "Ref.", "No.", "Dr.", "Mr.", "Ms.", "Prof.",
)


def _split_sentences(text: str) -> list[str]:
    """
    Split prose into sentences, protecting common abbreviations so their
    trailing period is not mistaken for a sentence boundary.
    """
    protected = text
    for ab in _SENTENCE_ABBREV:
        protected = protected.replace(ab, ab.replace(".", "\x00"))
    parts = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", protected)
    return [p.replace("\x00", ".").strip() for p in parts if p.strip()]


def _format_prose_as_bullets(text: str, max_bullets: int = 6) -> list[str]:
    """
    Turn a free-text paragraph into scannable Markdown.

    - A single sentence is rendered as a block-quote line.
    - Multiple sentences become a bullet list (one bullet per sentence),
      capped at `max_bullets`.
    """
    text = " ".join(str(text).split())   # collapse whitespace
    if not text:
        return []
    sentences = _split_sentences(text)
    if len(sentences) <= 1:
        return [f"> {text}"]
    bullets = [f"- {s}" for s in sentences[:max_bullets]]
    if len(sentences) > max_bullets:
        # fold any overflow into the last bullet so no content is lost
        remainder = " ".join(sentences[max_bullets:])
        bullets[-1] = bullets[-1] + " " + remainder
    return bullets


def _mermaid_label(text: str, max_len: int = 48) -> str:
    """
    Sanitise a string for use inside a Mermaid node label `id["..."]`.

    Mermaid breaks on raw double-quotes and a few other characters even
    inside quoted labels, so we replace/strip the problematic ones and
    truncate long labels to keep the figure readable.
    """
    t = " ".join(str(text).split())          # collapse whitespace/newlines
    t = t.replace('"', "'").replace("`", "'")
    t = t.replace("[", "(").replace("]", ")")
    t = t.replace("{", "(").replace("}", ")")
    t = t.replace("|", "/").replace("#", "no.")
    if len(t) > max_len:
        t = t[: max_len - 1].rstrip() + "…"
    return t or "—"


def _taxonomy_mermaid(
    root_label: str,
    top_level: list[str],
    second_level: dict[str, list[str]],
    anchor: str,
    max_top: int = 6,
    max_sub: int = 4,
) -> list[str]:
    """
    Render a survey's taxonomy as a Mermaid flowchart (left-to-right tree).

    Produces an inline ```mermaid code block that renders as a real figure
    on GitHub, VS Code, Obsidian, and most Markdown viewers — replacing the
    old ASCII text-art tree.  `anchor` makes node IDs unique per paper card
    so multiple diagrams on one page never collide.
    """
    # Node-id prefix unique to this card (strip non-alphanumerics from anchor)
    pid = "t" + re.sub(r"[^a-zA-Z0-9]", "", anchor)[:12]

    out: list[str] = ["```mermaid", "graph LR"]
    root_id = f"{pid}root"
    out.append(f'    {root_id}["{_mermaid_label(root_label)}"]')

    for i, cat in enumerate(top_level[:max_top]):
        cat_id = f"{pid}c{i}"
        out.append(f'    {cat_id}["{_mermaid_label(cat)}"]')
        out.append(f"    {root_id} --> {cat_id}")
        for j, sub in enumerate(second_level.get(cat, [])[:max_sub]):
            sub_id = f"{pid}c{i}s{j}"
            out.append(f'    {sub_id}["{_mermaid_label(sub)}"]')
            out.append(f"    {cat_id} --> {sub_id}")

    # Style: highlight the root node so the survey title stands out
    out.append(f"    style {root_id} fill:#dbeafe,stroke:#2563eb,stroke-width:2px")
    out.append("```")
    return out


def _html_attr(s: str) -> str:
    """Escape a string for use inside a double-quoted HTML attribute."""
    return (s.replace("&", "&amp;").replace("<", "&lt;")
             .replace(">", "&gt;").replace('"', "&quot;"))


def _taxonomy_tree_data(
    root_label: str,
    top_level: list[str],
    second_level: dict[str, list[str]],
    glossary: dict[str, str] | None = None,
    max_top: int = 8,
    max_sub: int = 6,
) -> dict:
    """
    A survey's taxonomy as nested {label, children, desc} for the interactive
    tree. `desc` (from the LLM glossary) is the plain-language "what this is"
    explanation shown in the click popup; matched to node names case-insensitively.
    """
    gloss = {str(k).strip().lower(): str(v) for k, v in (glossary or {}).items()}

    def desc_for(name: str) -> str:
        return gloss.get(str(name).strip().lower(), "")

    cats: list[dict] = []
    for cat in top_level[:max_top]:
        subs = []
        for s in (second_level.get(cat, []) or [])[:max_sub]:
            sub: dict = {"label": str(s)}
            d = desc_for(s)
            if d:
                sub["desc"] = d
            subs.append(sub)
        node: dict = {"label": str(cat)}
        d = desc_for(cat)
        if d:
            node["desc"] = d
        if subs:
            node["children"] = subs
        cats.append(node)
    return {"label": root_label, "children": cats}
